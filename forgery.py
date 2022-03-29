#######################################################################
# parse Excel Worksheet for correct SIC flag
# LOGIC FOR FORGARY & UTTER
# crime must match and must be convicted for it
# Parsing Reports Column (H) and writing results in column 15 ()
# (c) 2022 Theys Radmann
# ver 1.0 commited 20222-01-18
# final version 1.3 2022-01-02 
# Notes: TODO refine logic with client
#######################################################################
# modules/libararies needed
from openpyxl import Workbook, load_workbook
import re  # regex
import sys
import argparse
from common import ExcelFile, ExcelHeader, RegexSearch, Logger
# definition of crime categories
# order of some crimes in list is important for logic and efficiency
# first crime found and convicted for, ends check for further crimes, that's why
# put most frequent ones first 
ver = '1.3'
program_name = 'forgery'
logfile = 'forgery.log'
crimes = [
    r"forging",
    r"uttering",
    r"forgery",
    r"falsification[.,]",
    r"forge",
    r"counterfeiting[.]",
    r"(fake|altered|altering|false|(counterfeit(ing)?)|ficticious)( private| public| identity| client)? document[s]?",
    r"(fake|altered|altering|false|(counterfeit(ing)?)) report[s]?",
    r"counterfeit of( a)?( private| public| offical)? document[s]?",
    r"counterfeit ((bill[s]?)|bank receipts|finger prints|value added tax|travel authority)",
    r"fake( .+?){0,3} ((certificate[s]?)|(document[s]?))",
    r"(fake|false) papers",
    r"(fake|false|(counterfeit(ing)?))( of)?(( credit[s]?)| debit| prepaid)? card[s]?",
    r"(fake|false) (debit|credit) and (credit|debit) card[s]?",
    r"(counterfeit|fake)( .+?)? licen[sc]e[s]?",
    r"altering a(n)?",
    r"altering (land ownership|(record[s]?)|vehicle|school)",
    r"altering( .+?){0,4}? ((account[s]?)|(reimbursement[s]?)|receipts|forms|(record[s]?))",
    r"fake commercial private instrument",
    r"counterfeiting certificates",
    r"(counterfeit|fake|false)( .+?){0,3} ticket[s]?",
    r"counterfeit health insurance",
    r"fake( .+?){0,2} health insurance",
    r"(false|ficticious|fake) invoice[s]?",
    r"fake electricity bills",
    r"false ((record[s]?)|subscription)",
    r"false ID",
    r"false issuance of special VAT invoices",
    r"(fake|false) travel document[s]?",
    r"(false|fake|counterfeit)( .+?){0,2} identit(y|ies|ification)",
    r"fake( .+?)? visa",
    r"(fake|false|counterfeit|counterfeiting|tampered)( .+?){0,2} passport",
    r"(fake|counterfeit|counterfeiting) ATM",
    r"false expense",
    r"fake medical (job|(report[s]?))",
    r"(false|fake) (vehicle|car)",
    r"(altering|destroying|removing)( a)? vehicle identification number",
    r"fake bank (guarantee|accounts|cards)",
    r"fake bearer cheques",
    r"fake( payments)? receipt[s]?",
    r"(fake|(false(ly)?))( GST)? billing",
    r"(fake|false) credentials",
    r"fake government (identity|identification|papers|(document[s]?))",
    r"fake gst billing racket",
    r"(fake|false)( or fraudulent)? insurance",
    r"counterfeiting document[s]?",
    r"false data",
    r"false( .+?){0,4} ((document[s]?)|(report[s]?))",
    r"false (invoicing|(claim[s]?)|logbook|names)",
    r"false( print)? receipts",
    r"counterfeit printing",
    r"tampering with a public record",
    r"(fake|counterfeit|counterfeiting)( .+?){0,4} stamp[s]?",
    r"(counterfeit|false|fake)( .+?){0,2} (identification|identity)",
    r"((counterfeit(ing)*)|fake)( .+?){0,3} residence card[s]?",
    r"fraudulent( identification)? document[s]?",
    r"fraudulent income tax",
    r"(false|fake) documentation",
    r"false instrument[s]?",
    r"false( account)? statement[s]?",
    r"misleading information",
    r"false(( and| or) misleading)? ((document[s]?)|information|(statement[s]?))",
    r"false( and| or) misleading promotional material*",
    r"false( and| or) incomplete document[s]?",
    r"false(, )? inaccurate or incomplete documents",
    r"false income document[s]?",
    r"false (record|(report[s]?))",
    r"false social security number",
    r"(false|fake) tax (refund|return)",
    r"false federal income tax",
    r"false( and fictitious)? securities",
    r"(false|fake) signature[s]?",
    r"false account statements",
    r"false social security number",
    r"falsifying( .+?){0,4}(signatures|form)",
    r"falsifying ((document[s]?)|books)",
    r"fake ((account[s]?)|(degree[s]?)|demand|(joining letters)|emails|loan|lottery|letterheads|titles|(bill[s]?))",
    r"fake (lottery|registration|(marks card)|(sell agreement)|term deposit|(bail bond))",
    r"(fake|false|altering) ((land document[s]?)|titles)",
    r"(fake|false) (((para)?medical)|affidavit)",
    r"fake navy bills",
    r"illegal production of identification documents",
    r"fake receipts scam",
    r"fake admission[s]?",
    r"fake agreement",
    r"fake job ((certificate[s]?)|racket)",
    r"false annual compliance certification",
    r"answering firm's annual comliance questionnary falsely",
    r"applications that had been falsified",
    r"fake education(al)? ((certificate[s]?)|(documment[s]*))",
    r"false (testimony|writing)",
    r"fake( and fictitious)? auction vouchers",
    r"(false|counterfeit) value added tax",
    r"falsely issuing a special VAT",
    r"fake e-pass",
    r"fake examination[- ]certificate racket",
    r"fake army appointment letter",
    r"(false|fake)( .+?){0,3}? licen[cs]e[s]*",
    r"fictitious auto loan",
    r"ficticious prime bank securities",
    r"counterfeit(ing)?( .+?){0,2}? licen[cs]e[s]?",
    r"counterfeit(ing)? credit card[s]*",
    r"counterfeit (certificates|stamps)",
    r"counterfeiting of subpoenas",
    r"false entr(y|ies)",
    r"falsehood in( public)? document",
    r"(false|fake)( \w)? document[s]?",
    r"falsely (billing|(seeking (reimbursement[s]?))|signing|submitting)",
    r"falsely certif(ied|ying)",
    r"(falsely|fraudulently) inflat(e|ed|ing)",
    r"falsely (making out|issuing)( .+?)? value[ -]added[ -]tax (invoices|receipts)",
    r"falsely (making out|issuing)( .+?){0,4}? invoices",
    r"falsely reflect conversations with customers",
    r"falsely claiming (assistance|survivor benefits|reimbursement)",
    r"falsely (obtaining|receiving|recording|(report(ing)?)|(represent(ing|ed)))",
    r"falsely reflect conversations",
    r"falsely stated",
    "falsely completing (documments|forms)",
    r"falsely declaring( business)? expenses",
    r"falsely stating( .+?){1,2}? citizenship",
    r"false requisitions",
    r"false claims for expenses",
    r"false security services timesheets",
    r"false firm's documents",
    r"fictitious procurement contracts",
    r"fictitious( .+?){0,3}? ((statement[s]?)|(document[s]?)|(report[s]?)|(reimbursement[s]?)|insurance|disbursement)",
    r"forged ((instrument[s]?)|(document[s]?))",
    r"false and ((incomplete)|(misleading)) document[s]?",
    r"(inaccurate|incomplete) documents and statements",
    r"falsely seeking reimbursement",
    r"falsely signing( .+?){0,2}? customer names",
    r"falsely submitting personal expenses",
    r"falsely named as beneficiary",
    r"(false|misleading) information",
    r"counterfeit of a private document",
    r"counterfeit(ing)? of official",
    r"counterfeit( of government)? lottery",
    r"(counterfeit|fake) mark[ -]?sheet",
    r"(document|record) fals\w+",
    r"(document|record) forg\w+",
    r"unsworn falsfication",
    r"fabricating( .+){0,2} (business|financial|insurance|costs)",
    r"fabricat(ed|ing) charitable contributions deductions",
    r"fabricating a graft report",
    r"fabricating an account document",
    r"fabricating arbirtation documents",
    r"inaccurate documents and statements",
    r"improper alterations and other falsifications",
    r"(fabricated|fabricating)(. +?){0,4}? ((document[s]?)|((proposer )?information)|(report[s]?)|expense|reimbursement)",
    r"fabricating( total)? profit",
    r"counterfeit(ing)?( .+?){0,4} document[s]?",
    r"fake land (documents|papers|letters|title)",
    r"falsification of( .+?){0,2}? (official|certificates|documents|(evidence[s]?)|(work(er permit applications)?))",
    r"falsification of( .+?){0,2}? (expense|claims|reimbursement)",
    r"falsification of (funeral contracts|official statements in court|registered capital)",
    r"falsification of (disater relief vouchers|value added tax)",
    r"falsification of( .+?){0,2}? (identity|(public and private documents))",
    r"forg(ing|ed)( .+?){0,3}? (account|(invoice[s]?))",
    r"credit card counterfeiting",
    r"(document|ideological) falsehood",
    r"applications that had been falsified",
    r"answering firm's annual compliance questionnaire falsely",
    r"fake (Aadha(a)?r|Ayushman) card[s]?",
    r"fake pattadar passbooks",
    r"manipulation of (public procurement|land records|of the shares)"
    r"falsif((y(ing)?)|ied) (timesheets|paperwork|data submitted|evidence to prove own regitration|tax receipts)",
    r"falsif((y(ing)?)|ied) meeting minutes for procurement of monies",
    r"falsif((y(ing)?)|ied) (receipts for false claims|return-to-work data)",
    r"(false|inflated|inflating)( .+?){0,4}? ((reimbursement[s]?)|(billing{s]?)|(expense[s]?))",
    r"falsif((y(ing)?)|ied)( .+?){0,1}? expense reimbursement",
    r"falsif((y(ing)?)|ied)( .+?){0,4}? ((passport[s]?)|(document[s]?)|(record[s]?)|information|(form[s]?)|(letter[s]?))",
    r"falsif((y(ing)?)|ied)( .+?){0,4}? (travel|letters|(signature[s]?)|(affidavit[s]?)|insurance|statements|claims)",
    r"falsif((y(ing)?)|ied)( .+?){0,4}? (books|(report[s]?)|(authori[zs]ation)|(account[s]?)|funds|(certificate[s]?))",
    r"falsif((y(ing)?)|ied)( .+?){0,4}? (promissory notes|variable annuity|paperwork|federal tax income|resident cards)",
    r"falsif((y(ing)?)|ied)( .+?){0,2}? (wire transfer|data|results|annuity withdrawal|((customer(s)?)? application[s]?))",
    r"falsif((y(ing)?)|ied)( .+?){0,4}? (commision|account|cards|electoral|identity|loan applications)",
    r"falsif((y(ing)?)|ied)( .+?)? (((payroll|loan|mortgage) application(s)?)|((new )?account)|((sales|personal|income) tax returns))"
    ]
# below acquittal/dismissal triages should not be changed, unless new insight is found
aquittals = [
    r"ac?quitt(al|ed)",
    r"pardon(ed)?",
    r"case filed",
    r"dismissed",
    r"dropped"
]
dismissals = [
    r"dismiss(ed|al)",
    r"dropped",
    r"case filed"
]
words_apart = 20 # maximum distance of words apart from crime and conviction when matching cirme frst and conviction second
pre_conv = False
max_rep_length = 800
# functions
############################################################
def check_conviction(type, str_report, n):
    ''' Check if there was a convitcion for the offence type'''

# type : crime/offence (string/regex)
# str_report : record (report column) (string)
# r: row begin processed, for informational purposes only (debugging)
# returns 1 if found and issue follows conviction (no writes)
# returns 2 if found and issue is followed by conviction (no writes)
# returns -1 if issue is followed by conviction but too far apart (writes)
# returns -2 if conviction found with reversal (acquittal) found (no writes)
# returns 0 is no conviction was found at all (no writes)
# the reason for write case (-1) is in case there is no further processing that changes
# the status later. 
############################################################
    post_conv = 0
    long_flag = False
    global words_apart, Debug_Flg
    phrase = [
        r"found guilty",
        r"convicted",
        r"sentence[d]?",
        r"pleaded guilty",
        r"pleaded no contest",
        r"imprisoned",
        r"fined",
        r"arrested .+ serve",
        r"for conviction",
        r"ordered .*\s*to (pay|serve)",
        r"incarcerated",
        r"admitted guilt",
        r"served probation",
        r"to serve .* imprisonment",
        r'previous conviction[s]* .*?'
    ]
    # keywords must be near crime type if before conv
    # build search string with crime type
   
    for str in phrase:
        long_flag = False
        # search conviction before crime. Distance of words are checked here, but may not be necessary, 
        # as crime usually follows conviction after a few words if specified after.
        s_str = str + ' .*?' + type  # RegEx word followed by space and anythnig in between and the second word
        x = RegexSearch(s_str, str_report, n)
        if x: 
            # conviction found
            if Debug_Flg:
                log.output(n, x.group())
            # check if distance between offense and convictions is large
            words = re.split('\s', x.group())
            if len(words) > words_apart:
                # crime too far from sentence, look for another sentence further ahead, in case there are two
                n_idx = slice(x.start()+1, len(str_report)-1)
                y =RegexSearch(s_str, str_report[n_idx], n)
                if y:
                    words = re.split("\s", y.group())
                    if len(words) > words_apart:
                        print (n, "Too many words between issue and conv", end="\r")
                        post_conv = -1 # to flag for review
                        print(r, "SIC Review                            ", end='\r')
                        ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
                        if ListCheck:
                            ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: Remote connection between conviction and offence. From List")
                            xls.review += 1
                        else:
                            ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: Remote connection between conviction and offence")
                            xls.review += 1
                        return -1  # issue is too far for a conclusive conviction
                    # end if len()
                # enf if y
            # end if (len)
            # we have foud conviction, check for aquittals
            for tag in aquittals:
                s_str = type + '.*' + tag
                s_aquitt = RegexSearch(s_str, str_report, n)
                if s_aquitt:
                    print(n, "Dismissal found                                ", end='\r')   
                    return -2
                    # this may be revised
            # end for (aquitals)
            return 2
            
        # if not found, check the other way around (conviction after crime). 
        # Problem is, if there was a conviction for something different, in which
        # case we should not check for preious mentions of issues
        # this is difficult. Here some tries just to catch these common ones
        if "sentenced for" in str_report:
            continue
        if "pleaded guilty to" in str_report:
            continue
        if "found guilty for" in str_report:
            continue
        if "pleaded no contest to" in str_report:
            continue
        s_str = 'sentenced for \d+ years'
        x = RegexSearch(s_str, str_report, n)
        if not x:
            s_str = "sentence[d]* .*? *for "
            x = RegexSearch(s_str, str_report, n)
            if x:
                continue
        else:
            if Debug_Flg:
                print(type, "\n", str)
                print(n, "sentenced for x years found")
                input("enter")
        s_str = "sentence[d]* .*? *on charges of"
        x= RegexSearch(s_str, str_report, n)
        if x:
            continue
        s_str = "found guilty .*? *on charges of"
        x= RegexSearch(s_str, str_report, n)
        if x:
            continue
        s_str = 'pleaded guilty .*? *to'
        x= RegexSearch(s_str, str_report, n)
        if x:
            continue
        # now the check conviction after crime
        s_str = type + r'.*? ' + str
        x = RegexSearch(s_str, str_report, n)
        if x:
            if Debug_Flg:
                print(n, x.group())
            # found. if there are too many words between the type and the conviction phrase, assume review
            # there may be a later repetiion of the word (crime) wich decreases the word count, we check twice
            words = re.split("\s", x.group()) # split into words

            if len(words) > words_apart: # too many words in between, but there could be further mention of issue
                # look for issue further ahead
                n_idx = slice(x.start()+1, len(str_report)-1)
                y = RegexSearch(s_str, str_report[n_idx], n)
                if y:
                    words = re.split("\s", y.group())
                    if len(words) > words_apart:
                        print (n, "Too many words between issue and conv", end="\r")
                        post_conv = -1 # to flag for review
                        long_flag = True
                # end if (y)
            # end if (Len)
            if long_flag == False:
                for tag in aquittals:
                    s_str = type + '.*' + tag
                    s_aquitt = RegexSearch(s_str, str_report, n)
                    if s_aquitt:
                        print(n, "Aquittal found                                ", end='\r')   
                        return -2 
                    # this may be revised
                # end for (aquitals)
            # end if (long_flag)
            # found without aquittal
            return 1
        # end if (x)
    # end for (str)
    return post_conv
# end function ######################################################
#####################################################################
def check_item(item, str_Triage, r, pre_conv, src_text):
    '''Checks triage for record and writes approriate tag if offence/issue was found. '''
# normally called by check_issues()
# will not write to record if no match was found or exception (long content)
#####################################################################
    global Debug_Flg, ListCheck, ws, dismissals
    sic_tag = False
    s_crime = RegexSearch(item, str_Triage, r)
    if s_crime:
        
        if pre_conv:
            # check for dismissed
            for tag in dismissals:
                s_str = item + '.*' + tag # we may ommit item in search string
                s_diss = RegexSearch(s_str, str_Triage, r)
                if s_diss:
                    print(r, "Dismissal found                                ", end='\r')   
                    ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
                    ws.cell(row=r, column=head.col['Remarks'], value="Pre Conv: Dismissal found ["+src_text+"]")
                    xls.review += 1
                    return True
                # end if
            # end for
            print(r, 'SIC correct                                            ', end='\r')
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
            ws.cell(row=r, column=head.col['Remarks'], value="Pre Conv ["+src_text+"]")
            xls.sic_correct += 1
            return True
        # end if preconv
        # post conv processing
        chk = check_conviction(item, str_Triage, r)
        if chk == -1:
            # too far away, flag for review (for now). It was written in file in function
            print(r, "SIC Review                     ", end='\r')
            sic_tag = None
        if chk == 1:
            # write correct to sheet
            print(r, "SIC Correct                             ", end='\r')
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
            ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: Conviction after Triage ["+src_text+"]")
            xls.sic_correct += 1
            return True
        if chk == 2:
            # write correct to sheet
            print(r, "SIC Correct                             ", end='\r')
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
            ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: Conviction before Triage ["+src_text+"]")
            xls.sic_correct += 1
            return True
        if chk == -2:
            print(r, 'Review manually                                ', end='\r')
            ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
            ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: acquittal found ["+src_text+"]")
            xls.review += 1
            return True
        # if no conviction was found, check if there was a dismissal
        if chk == 0:
            for tag in dismissals:
                s_str = item + '.*' + tag # we may ommit item in search string
                s_diss = RegexSearch(s_str, str_Triage, r)
                if s_diss:
                    print(r, "Dismissal found                                ", end='\r')   
                    ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
                    ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: No conviction, dismissal found ["+src_text+"]")
                    xls.review += 1
                    return True # behaves like correct as no further offences are affeced if there is a dismissal
                    # this may be revised
                # end if
            # end for
            print (r, 'Tag correct - post conv', end='\r')
            ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
            ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: No conviction ["+src_text+"]")
            xls.review += 1
            return True

    # end if (s_crime true)
    return sic_tag  # can only be False or None here
        

#####################################################################
def check_issues(issues, str_Triage, r, preconv, Source):
# checks if crime was found and convicted
# # returns True if offense found
# (crime found and written in record in check_item()), 
# False (no tag/offense found), 
# and None (to review, long list or report, or distance) 
# 
#####################################################################
    global DebugFlg
    sic_tag = False
    # first, let's check review tags

    for x_crime in issues:
        # exclusions
        if x_crime == "uttering":
            x = RegexSearch(r"uttering[.].*( currency|( cheque[s]*)|( bank[ ]?note)| money)", str_Triage)
            if x:
                # counter triage, invalid
                continue
        if x_crime == "forge":
            x = RegexSearch(r"forge.*( currency|( cheque[s]*)|( bank[ ]?note)|( money))", str_Triage)
            if x:
                # invalid triage
                continue
        if x_crime == "forgery":
            x = RegexSearch(r"forgery[.,]?.*( currency|( cheque[s]*)|( bank[ ]?note)|( money))", str_Triage)
            if x:
                # 
                continue
        if x_crime == "forging":
            x = RegexSearch(r"forging[.,]?.*( currency|( cheque[s]*)|( bank[ ]?note)|( money))", str_Triage)
            if x:
                # 
                continue
        if x_crime == r"falsification[.,]":
            x = RegexSearch(r"falsification[.,].*(( currency)|( cheque[s]*)|( bank[ ]?note)|( money))", str_Triage)
            if x:
                continue
        # end exclusions ##########
        sic_tag = check_item(x_crime, str_Triage, r, preconv, Source)
        if sic_tag:
            break
    # end for (issues loop)
    return sic_tag
# end funcfions
###################################################

# start main program
print('Forgery & Uttering Processing\n' 'Version', ver)
DebugFlg = False
preconv_option = False

log = Logger(logfile)
xls = ExcelFile(program_name, ver)

if xls.debug_flag:
    DebugFlg = True
if xls.test_flag:
    row_limit = xls.row_limit
    Testflag = True
if xls.nolist:
    no_list = True
else:
    no_list = False
ws = xls.ws

    
head = ExcelHeader(ws)


if xls.preconv_option:
    print("Pre Conv mode")
    preconv_option = True


r = 0
print("Processing sheet")
for row in ws.rows:
    r += 1
    pre_conv = False
    if r == 1:
        continue # skip header (first row)
    # read row into variables (only useful ones)
    c_categories = ws.cell(row=r, column=head.col['Categories']).value       
    c_OfficialLists = ws.cell(row=r, column=head.col['OfficialLists']).value    
    c_AdditionalInfo = ws.cell(row=r,column=head.col['AdditionalInfo']).value
    c_Reports = ws.cell(row=r,column=head.col['Reports']).value     
    c_Type = ws.cell(row=r, column=head.col['Type']).value
    # c_Bio = ws.cell(row=r, column=head.col['Bio']).value
    c_status= ws.cell(row=r, column=head.col['Status']).value
    c_Triage =c_Reports
    # c_Bio = ws.cell(row=r, column=head.col['Bio']).value
    c_lists = [] # resets list of OifficialLists
    Extra = False
    ListCheck =False
    
    
    # Entities are flagged for manual review
    if c_Type == "E":
        # entity, flag for manual review
        print(r, "Entity: Review manually", end='\r')
        ws.cell(row=r, column=head.col['Status'], value="REVIEW MANNUALLY")
        ws.cell(row=r, column=head.col['Remarks'], value='ENTITY')
        xls.entities += 1
        continue
    
    if not c_Reports:
        ws.cell(row=r, column=head.col['Status'], value='NO REPORT')
        ws.cell(row=r, column=head.col['Remarks'], value='EMMPTY REPORT')
        print(r, "No report found.                         ", end='\r')
        xls.no_report += 1
        continue
    # check if in additional lists
    if c_OfficialLists and no_list == False:
        # extract lists from string
        # split string
        if DebugFlg:
            print(r, "List found")
        l_list = c_OfficialLists.split(';')
        i=0
        for tag in l_list:
            # look for tag in c_AdditionalInfo and extract string
            regex = '\['+tag+'\].*?\['
            if DebugFlg: 
                print (regex)
            x = RegexSearch(regex, c_AdditionalInfo, r)
            if x:
                c_lists.append(x.group())
                 # we do not need to strip the brackets
                print(r, "List match ", tag, "found", end="\r")
                i += 1
                has_list = True
            # end if
    #   # end for
        if DebugFlg:
            print(c_lists)
    # end if
    # we now have TagInfo populated
    # len(TagInfo) = mumber of elements (>0) or i , Ectra = True as flag, and i as the 
    # for str in TagInfo:
    #  we check crimes and convictions there as well at the end of the following loop
    if "CRIME" in c_categories:
        pre_conv = False
        xls.postconv +=1
    else:
        pre_conv = True
        xls.preconv +=1
        if preconv_option:
            # do not process crime records. Not implemented
            pass
    sic_tag = check_issues(crimes, c_Triage, r, pre_conv, 'RPT')
    # check for convvicted crimes in Report
    
    # now we check for Addidional List Tags
    if sic_tag == True:
        continue # go to next record
    if has_list:
        print("Checking additional in Lists", end="\r")
        ListCheck = True
        for x_Triage in c_lists:
            Long_Report = False
            if len(x_Triage) > max_rep_length:
                Long_Report = True
                ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
                ws.cell(row=r, column=head.col['Remarks'], value="LONG REPORT [LIST]")
                print(r, "Long list enry.                         ", end='\r')
                xls.long_entries += 1
                xls.review += 1
                continue # next list entry
            if sic_tag:
                break # no more list cheks
            sic_tag = check_issues(crimes, x_Triage, r, pre_conv, 'LST')
            if sic_tag == True:
                break
        # end for (list)
    # end if (extra/ lists)
    # if sic_tag was true, it was already written
    if sic_tag == True:
        continue
    
    if sic_tag == False:
        if Long_Report:
            print (r, "Review manually.                                 ", end='\r')
            ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
            ws.cell(row=r, column=head.col['Remarks'], value="Content to long.")
            xls.long_entries += 1
            xls.review += 1
            continue
        if pre_conv:
            print(r, "SIC incorrect                         ", end="\r")
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG INCORRECT")
            ws.cell(row=r, column=head.col['Remarks'], value='Pre Conv: No offense found')
            continue
        print(r, "SIC incorrect                              ", end='\r')
        if pre_conv == False:
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG INCORRECT")
            ws.cell(row=r, column=head.col['Remarks'], value='Post Conv: No offense found')
            xls.sic_incorrect += 1
        else:
            ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
            ws.cell(row=r, column=head.col['Ramarks'], value='Post Conv: No conviction found')
            xls.review += 1
    if sic_tag == None:
        print(r, "Review manually                            ", end='\r')
        ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
        ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: relation between crime and conviction not clear")
        xls.review += 1
    # end loop through rows
# write to new workbook

print('\nWriting and saving results spreadsheet ', xls.dest_file)

try:
    xls.ExcelSave()
except:
    input("\nCannot write to file. Try to close it first and press enter > ")
    print("Saving...")
    xls.ExcelSave()
print('Done')
log.toutput('Done')
# end program ######################################################################################################
 