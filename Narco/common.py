#######################################################################
# common functions/ classes for WordlCheck Triages
# ExcelHeader: readdsadrs header and assigns column numbers
#######################################################################
from openpyxl import Workbook
import sys
import re

##########################################################################
class ExcelHeader:
# reads header row and assigns column numbers
# called with ws (worksheet class)
# properties:
# col (type dict) col['Header_Name'] containes column number
# values are assigned at initialisation
########################################################################## 
    def __init__(self, ws):
        self.col = {
        'AdditionalInfo' : 0,
        'Categories' : 0,
        'OfficialLists' : 0,
        'Reports' : 0,
        'Status' : 0,
        'Type' : 0,
        'Remarks' : 0
        }
        # worksheet is pointer to open ws
        if not ws:
            # raise error
            print('ExcelHeader() : ws: No open or valid workheet')
            sys.exit()
        # read first line
        c = 0
        for c_col in ws.columns:
            c += 1
            if (ws.cell(row=1, column=c).value) == "Categories":
                self.col['Categories'] = c
            if (ws.cell(row=1, column=c).value) == "AdditionalInformation":
                self.col['AdditionalInfo'] = c
            if (ws.cell(row=1, column=c).value) == "OfficialLists":
                self.col['OfficialLists'] = c
            if (ws.cell(row=1, column=c).value) == "REPORTS":
                self.col['Reports'] = c
            if (ws.cell(row=1, column=c).value) == "TYPE":
                self.col['Type'] = c
            if (ws.cell(row=1, column=c).value) == "STATUS" or (ws.cell(row=1, column=c).value) == "Status":
                self.col['Status'] = c
        if self.col['Reports'] == 0:
            print('Reports column missing, check file')
            sys.exit()
        if self.col['AdditionalInfo'] == 0:
            print('AdditionalInformations column missing, check file')
            sys.exit()
        if self.col['Categories'] == 0:
            print('Reports column missing, check file')
            sys.exit()
        if self.col['Type'] == 0:
            print('Type column missing, check file')
            sys.exit()
        if self.col['Status'] == 0:
            self.col['Status'] = c + 1
            c += 1
            ws.cell(row=1, column=self.col['Status'], value='STATUS') # write missing status column
        self.col['Remarks'] = c + 1
        ws.cell(row=1, column=self.col['Remarks'], value = 'REMARKS')
    # end _init_
# end class ExcelHeader
############################################################################################
def RegexSearch(regex, String, r):
# helper function to search for regular expression to catch error
# and use case ignore flag
#############################################################################################
    try:
        p = re.compile(regex, re.IGNORECASE)
    except:
        print(r, 'Wrong regex:', regex)
        sys.exit()
    Mtch = p.search(String)
    return Mtch
# end function        