#######################################################################
# parse Excel Worksheet for correct SIC Tag
# LOGIC FOR NARCOTICS
# crime must match and must be convicted for it
# Parsing Reports Column (H) and writing results in column 15 ()
# (c) 2022 Theys Radmann
# ver 3.0, new triage for complete database
# version 4.0 totally new record processing
#######################################################################
# modules/libararies needed
from openpyxl import load_workbook, Workbook
import re  # regex
import sys
import argparse
from common import ExcelHeader, RegexSearch

# definition of crime categories
# order of some crimes in list is important for logic and efficiency
# first crime found and convicted for, ends check for further crimes, that's why
# put most frequent ones first 
ver = '3.0'
crimes = [
    r"racketeering and narcotics",
    r"involved in narcotics business",
    r"drugs with intent to distribute",
    r"((narcotics?)|(drugs?)) ((traffic(king)?)|distribution|import|transport|smuggling|violations?|sale|traffick)",
    r"((drugs?)|(narcotics?)) (delivery|(cultivat(e|ation))|(manufactur(e|ing))|(supply(ing)?)|charges|dealing|conspiracy|importation)",
    r"(drug|narcotics) (production|precursors|cultivation)",
    r"(narcotics|drugs) for sale",
    r"((narcotics?)|(drugs?)) ((charge[s]?)|racket|activities|offences?|burglary|convictions?|operations?|selling)",
    r"narcotics( .+?)? (charges|(crimes(s)?)|factory|felony|(for sale)|(offence(s)?))",
    r"((narcotics?)|(drugs?)) (production|precursors|cultivation)",
    r"(traffic|distribut|import|export|transport|smuggl|production|sell|deliver).*?((drugs?)|(narcotics?)|(controlled( dangerous)? substances?)|heroin|cocaine)",
    r"(distribut|cultivat|manufactur|dealing).*?(drugs?|narcotics?|cocaine|mari[jh]uana)",
    r"dealing in (drugs|nacrotics)",
    r"smuggl.* ((drugs?)|(narcotics?))",
    r"(sale|(dispens(e|ing))|delivery|(distribut(e|ion))) of( a)? (controlled substance|narcotics|drugs)",
    r"((distribute( a)?)|(dispense a)) controlled substances?"
    r"(conviction|charges) relat(ed|ing) to( a)? controlled substances?",
    r"(sale|robbery|theft|transfer) of (drugs|narcotics)",
    r"(selling|(steal(ing)?)|transfer) (drugs|narcotics|cocaine|crack|heroin)",
    r"sell( a)? (narcotics|controlled substances?|cocaine|hashish|heroin|mari[hj]uana)",
    r"posess and distribute (drugs|narcotics)?",
    r"posess(ing)? with( the)? intent to (distribute|supply) (drugs|narcotics)?",
    r"posessing( class a)? narcotics.*?(sales|supply)",
    r"racketeering involving (drugs|narcotics)",
    r"ditribut\w+(. +?)? ((drugs?)|narcotics)",
    r"production of (drugs|nartotics)",
    r"(drugs|narcotics)(. +?){0,3}? ((manufactur(e|ing))|conspiracy|dealing|supply|trade|trafficking)",
    r"(drugs|narcotics)(( .+){0,2}?[ -]related)? ((offence[s]?)|(crime[s]?)|(charge[s]?)|activities|convictions?|charges)",
    r"narcotics-trafficking activities",
    r"theft in nacrotics",
    r"intent on dealing",
    r"intent to distribute",
    r"purpose of trafficking",
    r"member of a narcotics synicate",
    r"violating( federal)? narcotics law",
    r"violating( the)? law(s)? of narcotics",
    r"(sale|robbery|theft|transfer) of (drugs|narcotics)",
    r"(selling|(steal(ing)?)|transfer) (drugs|Narcotics)",
    r"posession for the purpose of trafficking",
    r"narcotics for the purpose of trafficking",
    r"produc(e|ing)( .+?){0,2}? ((drugs?)|(narcotics?))",
    r"production of (drugs|narcotics?)",
    r"manufactur(e|ing)( of)?( a)? ((illicit substances?)|amphetamines)",
    r"(narcotics?|drugs?) posession.*?((deal(ing)?)|(sell(ing)?)|sale|(distribut(ion|e|ing))|(traffic(king?))|(cultivat(ion|e))|(supply(ing)?))",
    r"posession of (drugs|narcotics|controlled substances).*?((deal(ing)?)|sale|(sell(ing)?)|(distribut(ion|e|ing))|(supply(ing?))|traffick(ing)|(cultivat(e|ion)))",
    r"possesion of class a narcotics with intent to supply",
    r"(distribut|sell).*?(drugs|narcotics)",
    r"smuggl(e|ing|ed)( .+?){0,3}? ((drugs?)|(narcotics?))",
    r"suppl[y|ied|ying] (narcotics|drugs|amphetamine)"
    r"link between narcotic cartels",
    r"narcotics seized",
    r"seized narcotics",
    r"seizure of( .+){0,2} (narcotics?|drugs)"
    r"narcotics [(]hemp[)] trafficking",
    r"seizure of narcotics",
    r"distribute PCP",
    r"importing ecstacy",
    r"mantaining narcotics involved premises",
    r"(fixed|(import( a)?)) class a narcotics",
    r"distribution of pseudoephedrine",
    r"unlawful sale and promotion of prescription drugs",
    r"(smuggl|distribut|sell).*?(ketamine)",
    r"(smuggle|(supply(ing)?)) class [ab]",
    r"aggravated trafficking",
    r"dispensing of control(led)? substance(s)?",
    r"distribution of a listed chemical",
    r"narcotics[ -]trac[k]?fficking",
    r"(offences|felony) relat(ed|ing) to (drugs|narcotics|contrtolled substance)",
    r"conspiracy sale and posession of narcotics",
    r"(drugs|narcotic) conspiracy",
    r"running narcotics from",
    r"racketeering and narcotics",
    r"controlled substance - sell distribute",
    r"crimes against life and death"
]
# most common drugs

drugs = [
    r"cocaine( .+?)?",
    r"controlled substance[s]?",
    r"crack",
    r"ecstasy",
    r"mari[hj]uana",
    r"hydrocodone( .+?)?",
    r"isomethadone( .+?)?",
    r"narcotic[s]?",
    r"drug[s]?"
]
acquittals = [
    r"a[c]?quitt(al|ed)",
    r"pardon(ed)?",
    r"dismissed",
    r"dropped",
    r"case filed"
]

dismissals = [
    r"dismiss(ed|al)",
    r"dropped",
    r"case filed"
]
combinations = r"with( the)? intent to (deliver|ditribute|manufacture|sell|supply|traffic)"
    

words_apart = 30 # maximum distance of words apart from crime and conviction when matching cirme frst and conviction second
max_rep_length = 800 # maximum report length for processing, longer that this will get tagged for review
pre_conv_only = False
DebugFlg = False
# functions
####################################################################
def check_list_sic(list_tag, r):
#  returns true or false for trapping positive sic lists
####################################################################
# lists that trigger positive sic tag. could be read from a file
    TrueLists =[
        r"BDDNC",
        r"INNCB",
        r"NPNCB",
        r"NDLEA",
        r"PKANF",
        r"PHPDEA",
        r"RUFDCS",
        r"BPI-SDNT",
        r"BPI-SDNTK",
        r"SDNT",
        r"SDNTK",
        r"USINL",
        r"USSS-FRAA"
    ]
    list_status = [ False, "Null"]
    for str_list in TrueLists:
        if str_list in list_tag:
            list_status = [ True , str_list]
            return list_status
    return list_status 
# end check_sic_list()
############################################################
def check_conviction(c_crime, str_report, n):
# returning True, False, or None
# checks if there was a convitcion for the crime type
# c_crime : crime (string)
# str_report : record (report column) (string)
# r: row being processed, for informational purposes only (debugging)
# returns 1 if found and issue follows conviction (correct)
# returns 2 if found and issue is followed by conviction (correct)
# returns -1 if issue is folloed by conviction but too far apart (review manually)
# returns -2 if found but acquittal found as well (review manually)
# returns 0 is no conviction was found at all
# nothing gets written
# called from check_item() with post conv only
############################################################
    post_conv = 0
    long_flag = False
    global words_apart, DebugFlg
    phrase = [
        r"found guilty",
        r"convicted",
        r"sentence[d]?",
        r"pleaded guilty",
        r"pleaded no contest",
        r"imprisoned",
        r"fined",
        r"arrested .+ serve",
        r"for conviction",
        r"ordered .*\s*to (pay|serve)",
        r"incarcerated",
        r"admitted guilt",
        r"served probation",
        r"to serve .* imprisonment",
        r'previous conviction[s]* .*?'
    ]
    # keywords must be near crime type if before conv
    # build search string with crime type
   
    for conviction in phrase:
        long_flag = False
        # search keyword after conviction. Distance of words are not checked as crime usually follows conviction after a few words
        #  if specified after.
        s_str = conviction + ' .*?' + c_crime  # conviction before crime
        if DebugFlg:
            print(n, s_str)
            input("Press return ")
        x = RegexSearch(s_str, str_report, n)
        if x:
            post_conv = 2
            if DebugFlg:
                print(n, x.group())
            words = re.split('\s', x.group())
            if len(words) > words_apart:
                # crime too far from sentence, look for another sentence further ahead, in case there are two
                n_idx = slice(x.start()+1, len(str_report)-1)
                y = RegexSearch(s_str, str_report[n_idx], n)
                if y:
                    words = re.split("\s", y.group())
                    if len(words) > words_apart:
                        print (n, "Too many words between offense and conv", end="\r")
                        post_conv = -1 # to flag for review
                        # issue is too far for a conclusive conviction
                        # let's ignore this., but flag for review in case we do not find further evidence (return -1)
                        long_flag = True
                # issue is too far for a conclusive conviction
                # let's ignore this., but flag for review in case we do not find further evidence (return -1)
            # end if (len)
            # we have foud conviction, check for acquittals
            if long_flag == False:
                for tag in acquittals:
                    s_str = c_crime + '.*' + tag
                    s_acquitt = RegexSearch(s_str, str_report, n)
                    if s_acquitt:
                        print("Acquittal found                                ", end='\r')   
                        return -2
                        # this may be revised
                # end for (acquitals)
            # end if (long_flag)
            return post_conv
        # end if (x, conviction found)       
        # if not found, check the other way around. problem is if there was a conviction for somthing different, in which
        # case we should not check for previous mentions of issues and we should assume no conviction for this loop item
        # this is difficult. Here some tries just to catch these common ones
        if "pleaded guilty to" in str_report:
            continue
        if "found guilty for" in str_report:
            continue
        if "pleaded no contest to" in str_report:
            continue
        s_str = 'sentenced for \d+ years'
        x = RegexSearch(s_str, str_report, n)
        
        if not x:
            s_str = "sentence[d]* .*? *for "
            x = RegexSearch(s_str, str_report, n)
            if x:
                continue
        else:
            if DebugFlg:
                print(c_crime, "\n", str)
                print(n, "sentenced for x years found")
                input("enter")
        s_str = "sentence[d]* .*? *on charges of"
        x=  RegexSearch(s_str, str_report, n)
        if x:
            continue
        s_str = "found guilty .*? *on charges of"
        x= RegexSearch(s_str, str_report, n)
        if x:
            continue
        s_str = 'pleaded guilty .*? *to'
        x=  RegexSearch(s_str, str_report, n)
        if x:
            continue
        
        # now as previous were cleared, we check the other way around, conviction after crime
        s_str = c_crime + r'.*? ' + conviction
        x = RegexSearch(s_str, str_report, n)
        if x:
            post_conv = 1
            if DebugFlg:
                print(n, x.group())
            # found. if there are too many words between the type and the conviction phrase, assume review
            # there may be a later repetiion of the word wich decreases the word count, do we check twice
            words = re.split("\s", x.group()) # split into words

            if len(words) > words_apart: # too many words in between, but there could be further mention of issue
                # look for issue further ahead
                n_idx = slice(x.start()+1, len(str_report)-1)
                y = RegexSearch(s_str, str_report[n_idx], n)
                if y:
                    words = re.split("\s", y.group())
                    if len(words) > words_apart:
                        print (n, "Too many words between issue and conv", end="\r")
                        post_conv = -1 # to flag for review
                        long_flag = True
                else:
                    long_flag = True
                    post_conv = -1
                # end if (y)
            # end if (Len)
            if long_flag == False:
                for tag in acquittals:
                    s_str = c_crime + '.*' + tag
                    s_acquitt = RegexSearch(s_str, str_report, n)
                    if s_acquitt:
                        print(n, "Acquittal found                                ", end='\r')   
                        return -2 
                    # this may be revised
                # end for (aquitals)
            # end if (long_flag)
            return post_conv
        # end if (x)
    # end for (str)
    return post_conv
#####################################################################
def check_item(item, str_Triage, r, TrType, preconv_only):
# treturns True, False or None
# applies SIC logic to apply for one record
# checks if offence is found. If true, applies logic according to pre or post conv
# writes to record and returns true to mark end of further item search
# If False, returns returns false (no writing, as next item could be found)
# Returns none if offence was found but the match string is very long so
# it may refer to a different connection
#####################################################################
    global DebugFlg, ListCheck, ws, dismissals
    sic_tag = False
    if len(str_Triage) > 500:
        # mark as too long
        if preconv_only == False:
            ws.cell(row=r, column=head.col['Status'], value='REVIEW MANUALLY')
            ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: Entry too long")
            return True    
    s_offence = RegexSearch(item, str_Triage, r)
    if s_offence:
        if len(s_offence.group()) > 150:
            # should mark as review, as we found a remote connection. this is very rare
            # and will happen only with regex that include an open number of characters to match
            sic_tag = None
            return sic_tag
        if preconv_only == True:
            # we found offense, check dismissal
            for tag in dismissals:
                s_str = item + '.*' + tag # we may ommit item in search string
                s_diss = RegexSearch(s_str, str_Triage, r)
                if s_diss:
                    print(r, "Review, dismissal found                                ", end='\r')   
                    ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
                    ws.cell(row=r, column=head.col['Remarks'], value="Pre Conv: Dismissal found.")
                    return True # behaves like correct as no further offences are affeced if there is a dismissal
                # end if
            # end for
            # else, no dismissal found
            print(r, "SIC Correct                             ", end='\r')
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
            if ListCheck:
                ws.cell(row=r, column=head.col['Remarks'], value="Pre Conv (List) [: "+TrType+"]")
            else:
                ws.cell(row=r, column=head.col['Remarks'], value="Pre Conv ["+TrType+"]")
            return True
        # end if
        # post conv now     
        chk = check_conviction(item, str_Triage, r)
        if chk == -1:
            # too far away, flag for review (for now)
            print(r, "SIC Review                            ", end='\r')
            ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
            if ListCheck:
                ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: Remote connection between conviction and offence ["+TrType+"]. From List")
            else:
                ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: Remote connection between conviction and offence ["+TrType+"]")
            return True
        if chk == 1:
            # write correct to sheet
            print(r, "SIC Correct                             ", end='\r')
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
            if ListCheck:
                ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: Conviction after offence ["+TrType+"]. From List")
            else:
                ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: Conviction after offence ["+TrType+"]")
            return True
        if chk == 2:
            # write correct to sheet
            print(r, "SIC Correct                             ", end='\r')
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
            if ListCheck:
                ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: Conviction before offence ["+TrType+"]. From List")
            else:
                ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: Conviction before pffence ["+TrType+"].")
            return True
        if chk == -2:
            print(r, 'Review manually, acquittal found                            ', end='\r')
            ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
            if ListCheck:
                ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: conviction with acquittal ["+TrType+"]. From List")
            else:
                ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: conviction with acquittal ["+TrType+"].")
            return True
        if chk == 0:
            # no conviction found, look for dismissals
            for tag in dismissals:
                s_str = item + '.*' + tag # we may ommit item in search string
                s_diss = RegexSearch(s_str, str_Triage, r)
                if s_diss:
                    print(r, "Review, dismissal found                                ", end='\r')   
                    ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
                    ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: Dismissal found ["+TrType+"].")
                    return True # behaves like correct as no further offences are affeced if there is a dismissal
                # end if
            # end for
            print (r, 'Tag correct - pre conv', end='\r')
            ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
            if ListCheck:
                ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: No conviction ["+TrType+"]. From List")
            else:
                ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: No conviction ["+TrType+"].")
            return True
    # end if (crime found)
    return sic_tag
    
#####################################################################
def check_issues(issues, str_Triage, r, preconv):
# checks if crime was found and convicted
# # returns True (crime found and written in record), False, and None (to review) 
# writes record if correct
#####################################################################
    global DebugFlg, cws, NoChemCheck
    sic_tag = False
   
    # 1st pass: general
    for r_crime in issues:
        sic_tag = check_item(r_crime, str_Triage, r, 'general', preconv)
        if sic_tag:
            break
    if sic_tag:
        return True
        
    # 2nd pass, chemicals
    if NoChemCheck:
        return sic_tag
    n = 0
    for c_row in cws.rows:
        n += 1
        chemical = cws.cell(row=n, column=1).value
        sic_tag = check_item(chemical, str_Triage, r, 'chem', preconv)
        if sic_tag:
            break
    return sic_tag
# end functions
###################################################
#####################################################################
def check_pharma(str_Triage, r):
# checks if present in pharma list
# # returns True (found in pharma) or Fale (not found)  
# writes record if correct
#####################################################################
    global DebugFlg, pws, NoChemCheck
    sic_tag = False
 
    n = 0
    for p_row in pws.rows:
        n += 1
        chemical = pws.cell(row=n, column=1).value
        s_offence = RegexSearch(chemical, str_Triage, r)
        if s_offence:
            return True
    return sic_tag
# end functions
###################################################
# start program
Testflag = False
DebugFlg = False
TrueCondition = False
offence_found = False
RowLimit = 0
NoChemCheck = False
parser = argparse.ArgumentParser(description='Process Narcotics SIC', prog='nacrotics.py')
parser.add_argument("--pc", help="Chcek pre-convition only)", action='store_true')
parser.add_argument("--version",help="Displays version only", action='version', version='%(prog)s ' + ver)
parser.add_argument("--debug", help="Debug mode", action='store_true')
parser.add_argument('filename', help="filename to read")
parser.add_argument('-t', '--test', help='run for a limited number of rows', type=int)
parser.add_argument("--nolist", help="No chemicals list is checked", action='store_true')
args = parser.parse_args()
if args.debug:
    DebugFlg = True
if args.pc:
    preconv_only = True
    print("Pre conviction option")
if args.nolist:
    NoChemCheck = True
    print("Chemical file is not checked.")
if args.test:
    Testflag = True
    RowLimit = args.test   
    print ('Test: processing only', RowLimit, ' rows') 
org_file = args.filename
if ".xlsx" not in org_file:
    dest_file = org_file + ' Passed.xlsx'
    WSheet = org_file
    org_file = org_file + '.xlsx'
else:
    file_parts = org_file.split('.')
    dest_file = file_parts[0] + ' Passed.xlxs'
    WSheet = file_parts[0]
if DebugFlg:
    print('Org: ', org_file)
    print('Dest: ', dest_file)
    input('enter > ')

print( 'Loading spreadsheet', org_file)
# check if filename exists
#
try:     
    wb = load_workbook(filename=org_file)
except:
    print("cannot open file", org_file)
    sys.exit()
ws = wb[WSheet]
head = ExcelHeader(ws)
if DebugFlg:
 print(head.col)
 input('Enter > ')
 # load chemicals list
if not NoChemCheck:
    print('Loading chemicals list from file.')
    try:
        cwb = load_workbook('NarcList.xlsx')
    except:
        print('Coud not open file NarcList.xlsx. Is it open?')
        input('Enter > ')
        cwb = load_workbook('NarcList.xlsx')
    # load workheet for pharma
    cws = cwb['Sheet1']
    print('Loading Pharma List form file.')
    try:
        pwb = load_workbook('PharmList.xlsx')
    except:
        print('Coud not open file PharmList.xlsx. Is it open?')
        input('Enter > ')
        pwb = load_workbook('NarcList.xlsx')
    pws = pwb['Sheet1']

r = 0
print("Processing worksheet")
for g_row in ws.rows:
  
    pre_conv_only = False
    r += 1
    if r == 1:
        continue
    if Testflag:
        if r > RowLimit:
            break
    # read row into variables (only useful ones)
    c_categories = ws.cell(row=r, column=head.col['Categories']).value       
    c_OfficialLists = ws.cell(row=r, column=head.col['OfficialLists']).value    
    c_AdditionalInfo = ws.cell(row=r,column=head.col['AdditionalInfo']).value    
    c_Reports = ws.cell(row=r,column=head.col['Reports']).value     
    c_Type = ws.cell(row=r, column=head.col['Type']).value
    # c_Bio = ws.cell(row=r, column=head.col['Bio']).value
    c_status= ws.cell(row=r, column=head.col['Status']).value
    c_Triage = c_Reports
    c_lists = [] # resets list of OifficialLists
    ListsPresent = False
    LongReport = False
    ListCheck = False
    Combi = False
    

    # Entities are flagged for manual review
    # 1. Entities are flagged for review
    if c_Type == "E":
        # entity, flag for manual review
        print(r, "Entity: Review manually", end='\r')
        ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
        ws.cell(row=r, column=head.col['Remarks'], value='ENTITY')
        print(r, "Entity.                                 ", end='\r')
        continue
    if DebugFlg:
        print(r, c_Reports)
        input('Enter > ')
    
    
    if not c_Reports:
        ws.cell(row=r, column=head.col['Status'], value='NO REPORT')
        ws.cell(row=r, column=head.col['Remarks'], value='No report column')
        print(r, "No report found.                         ", end='\r')
        continue
    # Long reports
    if len(c_Reports) > max_rep_length:
        ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
        ws.cell(row=r, column=head.col['Remarks'], value='Report content too long')
        print(r, "Report too long.                         ", end='\r')
        continue
    # 2. check if in additional lists, and populate list content if found
    if c_OfficialLists:
        sic_list = check_list_sic(c_OfficialLists, r)
        if sic_list[0] == True:
            print(r, "Tagged list", end='\r')
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
            ws.cell(row=r, column=head.col['Remarks'], value='OFFICIAL LIST : ' + sic_list[1])
            continue # no further processing needed
        # Check if there are brackets for lists in AdditionalInfo and populate set
        # split string
        if DebugFlg:
            print(r, "List found")
        l_list = c_OfficialLists.split(';')
        i=0
        for tag in l_list:
            # look for tag in c_AdditionalInfo and ListsPresent string
            regex = r'\['+tag+r'\].*?\['
            #_DEBUG print (regex)
            x = RegexSearch(regex, c_AdditionalInfo, r)
            if x:
                c_lists.append(x.group())
                 # we do not need to strip the brackets
                print(r, "List match ", tag, "found            ", end="\r")
                i += 1
                ListsPresent = True
            # end if
        # end for
    # end if (OfficialLists)
    # 3. Check for Combinations in report and lists
    x_comb = RegexSearch(combinations, c_Triage, r)
    if x_comb: 
        print(r, "Combination found", end='\r')
        ws.cell(row=r, column=head.col['Status'], value='REVIEW MANUALLY')
        ws.cell(row=r, column=head.col['Remarks'], value='COMBINATION')
        continue
    # end if
    # Combinations in lists, we do not check for long reports now.
    if ListsPresent:
        for x_Triage in c_lists:
            x_comb = RegexSearch(combinations, x_Triage, r)
            if x_comb: 
                print(r, "Combination found", end='\r')
                ws.cell(row=r, column=head.col['Status'], value='REVIEW MANUALLY')
                ws.cell(row=r, column=head.col['Remarks'], value='COMBINATION (LIST)')
                Combi = True
                break
            # end if
        # end for        
    # end for
    # we end further processing if combination was found
    if Combi:
        continue            
    # we now have official lists poulated if matched
    # ver 3.1 post conv, if narcotics crime catagory, it is correct if no pharma entry was found. This is done before anything else
    if "CRIME - NARCOTICS" in c_categories:
        # run through pharma list
        if check_pharma(c_Triage, r):
            print(r, "Pharma found in narcotics crime                ", end='\r')
            ws.cell(row=r, column=head.col['Status'], value='REVIEW MANUALLY')
            ws.cell(row=r, column=head.col['Remarks'], value='Post Conv: Narcotics Crime Category, pharma found')
        else:
            # check in lists if present
            if ListsPresent:
                for x_Triage in c_lists:
                    sic_tag = check_pharma(x_Triage, r)
                    if sic_tag:
                        print(r, "Pharma found in narcotics crime", end='\r')
                        ws.cell(row=r, column=head.col['Status'], value='REVIEW MANUALLY')
                        ws.cell(row=r, column=head.col['Remarks'], value='Post Conv: Narcotics Crime Category, pharma found (LIST')
                        break
                    # end if
                # end for
                if not sic_tag:
                    ws.cell(row=r, column=head.col['Status'], value='SIC TAG CORRECT')
                    ws.cell(row=r, column=head.col['Remarks'], value='Post Conv: Narcotics Crime Category')
                # end if
            else:
                # no lists present
                ws.cell(row=r, column=head.col['Status'], value='SIC TAG CORRECT')
                ws.cell(row=r, column=head.col['Remarks'], value='Post Conv: Narcotics Crime Category')
            # end if list        
        # end if check pharma
        continue        
    # end if    
    # 4. Chcek tags for logic. Pre conv or post conv depending on record
    # flg pre or pos conv
    if "CRIME" in c_categories:
        pre_conv_only = False
    else:
        pre_conv_only = True
    # 
    # we pass Chemicals, and General trhough Reports and lists. 
        
    sic_tag = check_issues(crimes, c_Triage, r, pre_conv_only)
    # now we check for Addidional List Tags
    if sic_tag == True:
        continue # go to next record
    # if not found yet, continue checking in lists if present
    if ListsPresent:
        print(r, "Checking additional in Lists", end="\r")
        LongReport = False
        ListCheck = True
        for x_Triage in c_lists:
            # bwloe is checked in check_item()
            #if len(x_Triage) > 500:
            #    LongReport = True
            #    ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
            #    ws.cell(row=r, column=head.col['Remarks'], value="Content of list entry too long.")
            #    print(r, "Long list enry.                         ", end='\r')
            #    continue # check next list tag, if found successfull, the record will be overwritten
            sic_tag = check_issues(crimes, x_Triage, r, pre_conv_only)
            if sic_tag == True:
                break
        # end for
        # if long report was detected and no crime found
    # end if
    # if sic_tag was true, it was already written
    if sic_tag == True:
        continue
    if sic_tag == False:
        if LongReport:
            print (r, "Review manually.                                 ", end='\r')
            ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
            ws.cell(row=r, column=head.col['Remarks'], value="List Content to long.")
            continue
        if pre_conv_only:
            print(r, 'SIC incorrect                                   ', end='\r')
            ws.cell(row=r, column=head.col['Status'], value="TAG SHOULD BE REMOVED")
            ws.cell(row=r, column=head.col['Remarks'], value="Pre Conv: No offence found")
            continue
        print(r, "SIC incorrect                              ", end='\r')
        ws.cell(row=r, column=head.col['Status'], value="TAG SHOULD BE REMOVED")
        ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: No offence found")
    if sic_tag == None:
        print(r, "Review manually. Long text.                            ", end='\r')
        ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
        ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: relation between crime and conviction not clear")
    # end loop through rows
# write to new workbook
print('\nWriting and saving results in file', dest_file, '...' )
try:
    wb.save(dest_file)
except:
    input("\nCannot write to file. Try to close it first and press enter > ")
    print("Saving...")
    wb.save(dest_file)
print('Done')
# end program ######################################################################################################
 