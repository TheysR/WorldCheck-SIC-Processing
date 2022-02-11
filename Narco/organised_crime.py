#######################################################################
# parse Excel Worksheet for correct SIC flag
# LOGIC FOR ORGANISED CRIME
# crime must match and must be convicted for it
# (c) 2022 Theys Radmann
# ver 1.0, initial version
#######################################################################
# modules/libararies needed
from weakref import WeakSet
from mysqlx import Row
from openpyxl import load_workbook, Workbook
import re  # regex
import sys
import argparse
from common import ExcelHeader
# definition of crime categories
# order of some crimes in list is important for logic and efficiency
# first crime found and convicted for, ends check for further crimes, that's why
# put most frequent ones first 
ver = '3.0'
# we read crimes from file

# manually review tags
crimes = []
RevMan = [
    r"Collaborator of Los",
    r"conspiracy ring",
    r"cutting agent conspiracy",
    r"(distribution|growing|narcotics|smuggling|trafficking) conspiracy",
    r"growing organisation",
    r"Member of Los",
    r"Militia",
    r"trade network",
    r"trafficking (\group",
    r"fixing syndicate",
    r"production group",
    r"(\FARC\)",
    r"(\AUC\)",
    r"terrorism",
    r"Revolutionary Armed Forces of Colombia",
    r"ETA",
    r"terror related",
    r"terror"

]
# most common drugs

aquittals = [
    r"aquitt(al|ed)",
    r"pardon(ed)?"
]
dismissals = [
    r"dismiss(ed|al)",
    r"dropped",
    r"case filed"
]

words_apart = 20 # maximum distance of words apart from crime and conviction when matching cirme frst and conviction second
pre_conv = False
DebugFlg = False
# functions
############################################################
def check_conviction(type, str_report, n):
# returning True, False, or None
# checks if there was a convitcion for the crime type
# type : crime (string)
# str_report : record (report column) (string)
# r: row begin processed, for informational purposes only (debugging)
# returns 1 if found and issue follows conviction
# returns 2 if found and issue is followed by conviction
# returns -1 if issue is folloed by conviction but too far apart
# returns 0 is no conviction was found at all
############################################################
    post_conv = -1
    long_flag = False
    global words_apart, DebugFlg, preconv_option
    phrase = [
        r"found guilty",
        r"convicted",
        r"sentence[d]*",
        r"pleaded guilty",
        r"pleaded no contest",
        r"imprisoned",
        r"fined",
        r"arrested .+ serve",
        r"for conviction",
        r"ordered .*\s*to (pay|serve)",
        r"incarcerated",
        r"admitted guilt",
        r"served probation",
        r"to serve .* imprisonment",
        r'previous conviction[s]* .*?'
    ]
    # keywords must be near crime type if before conv
    # build search string with crime type
   
    for str in phrase:
        long_flag = False
        # search keyword after conviction. Distance of words are not checked as crime usually follows conviction after a few words
        #  if specified after.
        s_str = str + ' .*?' + type  # RegEx word followed by space and anythnig in between and the second word
        if DebugFlg:
            print(n, s_str)
            input("Press return ")
        try:
            p = re.compile(s_str, re.IGNORECASE)
        except:
            print('Wrong regex:', s_str)
            sys.exit()
        x = p.search(str_report)
        if x:
            if DebugFlg:
                print(n, x.group())
            words = re.split('\s', x.group())
            if len(words) > words_apart:
                # crime too far from sentence, look for another sentence further ahead, in case there are two
                y =p.search(str_report, x.start()+1)
                if y:
                    words = re.split("\s", y.group())
                    if len(words) > words_apart:
                        print (n, "Too many words between issue and conv", end="\r")
                        post_conv = -1 # to flag for review
                        # issue is too far for a conclusive conviction
                        # let's ignore this., but flag for review in case we do not find further evidence (return -1)
                        long_flag = True
                # issue is too far for a conclusive conviction
                # let's ignore this., but flag for review in case we do not find further evidence (return -1)
            # end if (len)
            # we have foud conviction, check for aquittals
            for tag in aquittals:
                s_str = type + '.*' + tag # may be let's not check for crime type as well
                q = re.compile(s_str)
                s_aquitt = q.search(s_str)
                if s_aquitt:
                    print("Dismissal found                                ", end='\r')   
                    return -2 # although not coreect, is behaves like correct as no further offences are affeced if there is a dismissal
                    # this may be revised
            # end for (aquitals)
            return 2
            
        # if not found, check the other way around. problem is if there was a conviction for somthing different, in which
        # case we should not check for preious mentions of issues
        # this is difficult. Here some tries just to catch these common ones
        if "sentenced for" in str_report:
            continue
        if "pleaded guilty to" in str_report:
            continue
        if "found guilty for" in str_report:
            continue
        if "pleaded no contest to" in str_report:
            continue
        s_str = 'sentenced for \d+ years'
        p = re.compile(s_str, re.I)
        x = p.search(str_report)
        if not x:
            s_str = "sentence[d]* .*? *for "
            p = re.compile(s_str, re.I)
            x = p.search(str_report)
            if x:
                continue
        else:
            if DebugFlg:
                print(type, "\n", str)
                print(n, "sentenced for x years found")
                input("enter")
        s_str = "sentence[d]* .*? *on charges of"
        p= re.compile(s_str, re.I)
        x= p.search(str_report)
        if x:
            continue
        s_str = "found guilty .*? *on charges of"
        p= re.compile(s_str, re.I)
        x= p.search(str_report)
        if x:
            continue
        s_str = 'pleaded guilty .*? *to'
        p= re.compile(s_str, re.I)
        x= p.search(str_report)
        if x:
            continue
         # now the other way around
        s_str = type + r'.*? ' + str
        p = re.compile(s_str, re.IGNORECASE)
        x = p.search(str_report)
        if x:
            if DebugFlg:
                print(n, x.group())
            # found. if there are too many words between the type and the conviction phrase, assume review
            # there may be a later repetiion of the word wich decreases the word count, do we check twice
            words = re.split("\s", x.group()) # split into words

            if len(words) > words_apart: # too many words in between, but there could be further mention of issue
                # look for issue further ahead
                y = p.search(str_report, x.start()+1)
                if y:
                    words = re.split("\s", y.group())
                    if len(words) > words_apart:
                        print (n, "Too many words between issue and conv", end="\r")
                        post_conv = -1 # to flag for review
                        long_flag = True
                # end if (y)
            # end if (Len)
            if long_flag == False:
                for tag in aquittals:
                    s_str = type + '.*' + tag
                    q = re.compile(s_str)
                    s_aquitt = q.search(s_str)
                    if s_aquitt:
                        print(n, "Aquittal found                                ", end='\r')   
                        return -2 
                    # this may be revised
                # end for (aquitals)
            # end if (long_flag)
            # found without aquittal
            return 1
        # end if (x)
    # end for (str)
    return post_conv
#####################################################################
def check_item(item, str_Triage, r):
#####################################################################
    global pre_conv, preconv_option, DebugFlg, ListCheck, ws, dismissals
    sic_crime = False
    try:
        p = re.compile(item, re.I)
    except:
        print(r, 'Regex error:', item)
        sys.exit()
    s_crime = p.search(str_Triage)
    if s_crime:
        pre_conv = True
        if len(s_crime.group()) > 100:
            # should mark as review, as we found a remote connection
            sic_crime = None
            return sic_crime
        if preconv_option:
            for tag in dismissals:
                s_str = item + '.*' + tag # we may ommit item in search string
                q = re.compile(s_str)
                s_diss = q.search(str_Triage)
                if s_diss:
                    print(r, "Dismissal found                                ", end='\r')   
                    ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
                    ws.cell(row=r, column=head.col['Remarks'], value="Offence found but dismissal found.")
                    return True # although not coreect, is behaves like correct as no further offences are affeced ig there is a dismissal
                    # this may be revised
                # end if
            # end for    
            print(r, "SIC Correct                                ", end='\r')
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
            if ListCheck:
                ws.cell(row=r, column=head.col['Remarks'], value="From LIst")
            return True
        # end if (preconv_option)
        chk = check_conviction(item, str_Triage, r)
        if chk == -1:
            # too far away, flag for review (for now)
            print(r, "SIC Review                     ", end='\r')
            sic_crime = None
        if chk == 1:
            # write correct to sheet
            print(r, "SIC Correct                             ", end='\r')
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
            if ListCheck:
                ws.cell(row=r, column=head.col['Remarks'], value="Conviction after Triage. From List")
            else:
                ws.cell(row=r, column=head.col['Remarks'], value="Conviction after Tag")
            return True
        if chk == 2:
            # write correct to sheet
            print(r, "SIC Correct                             ", end='\r')
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
            if ListCheck:
                ws.cell(row=r, column=head.col['Remarks'], value="Conviction before Triage. From List")
            else:
                ws.cell(row=r, column=head.col['Remarks'], value="Conviction before Tag")
            return True
        if chk == -2:
            print(r, 'Review manually                                ', end='\r')
            ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
            if ListCheck:
                ws.cell(row=r, column=head.col['Remarks'], value="Offence with conviction but with aquittal. From List")
            else:
                ws.cell(row=r, column=head.col['Remarks'], value="Offence with conviction but with aquittal")
            return True
    return sic_crime
    # end if (s_crime true)    

#####################################################################
def check_issues(issues, str_Triage, r):
# checks if crime was found and convicted
# # returns True (crime found and written in record), False, and None (to review) 
# writes record if correct
#####################################################################
    global pre_conv, DebugFlg, pws
    sic_crime = False

    for x_crime in issues:
        sic_crime = check_item(x_crime, str_Triage, r)
        if sic_crime:
            break
    # end for (issues loop)
    return sic_crime
# end functions
###################################################

# start program
Testflag = False
DebugFlg = False
preconv_option = False
TrueCondition = False
offence_found = False
RowLimit = 0
parser = argparse.ArgumentParser(description='Process Narcotics SIC', prog='nacrotics.py')
parser.add_argument("--pc", help="Chcek pre-convition only)", action='store_true')
parser.add_argument("--version",help="Displays version only", action='version', version='%(prog)s ' + ver)
parser.add_argument("--debug", help="Debug mode", action='store_true')
parser.add_argument('filename', help="filename to read")
parser.add_argument('-t', '--test', help='run for a limited number of rows', type=int)
args = parser.parse_args()
if args.debug:
    DebugFlg = True
if args.pc:
    preconv_option = True
    print("Pre conviction option")
if args.test:
    Testflag = True
    RowLimit = args.test   
    print ('Test: processing only', RowLimit, ' rows') 
org_file = args.filename
if ".xlsx" not in org_file:
    if preconv_option:
        dest_file = org_file + ' Preconv Passed.xlsx'
    else:
        dest_file = org_file + ' Passed.xlsx'
    WSheet = org_file
    org_file = org_file + '.xlsx'
else:
    file_parts = org_file.split('.')
    if preconv_option:
        dest_file = file_parts[0] + ' Preconv Passed.xlxs'
    else:
        dest_file = file_parts[0] + ' Passed.xlxs'
    WSheet = file_parts[0]
if DebugFlg:
    print('Org: ', org_file)
    print('Dest: ', dest_file)
    input('enter > ')

print( 'Loading spreadsheet', org_file)
# check if filename exists
#
try:     
    wb = load_workbook(filename=org_file)
except:
    print("cannot open file", org_file)
    sys.exit()
ws = wb[WSheet]
head = ExcelHeader(ws)
if DebugFlg:
 print(head.col)
 input('Enter > ')
 # loaf chemicals list
print('Loading chemnicals list from file.')
try:
    cwb = load_workbook('OrgCrimes.xlsx')
except:
    print('Coud not open file OrgCrimes.xlsx. Is it open?')
    input('Enter > ')
    cwb = load_workbook('NarcList.xlsx')
# load workheet for crimes
cws = cwb['Sheet1']
# populate crimes
r = 1
for row in cws.rows:
    crimes.append(cws.cell(row=r, column=1).value)
    r += 1
print("Processing worksheet")
r = 0
for row in ws.rows:
    pre_conv = False
    r += 1
    if r == 1:
        continue
    if Testflag:
        if r > RowLimit:
            break
    # read row into variables (only useful ones)
    c_categories = ws.cell(row=r, column=head.col['Categories']).value       
    c_OfficialLists = ws.cell(row=r, column=head.col['OfficialLists']).value    
    c_AdditionalInfo = ws.cell(row=r,column=head.col['AdditionalInfo']).value
    c_Bio = ws.cell(row=r, column=head.col['Bio']).value    
    c_Reports = ws.cell(row=r,column=head.col['Reports']).value     
    c_Type = ws.cell(row=r, column=head.col['Type']).value
    c_status= ws.cell(row=r, column=head.col['Status']).value
    c_Triage = c_Reports
    TagStr = [] # resets list of OifficialLists
    Extra = False
    LongReport = False
    ListCheck = False
    

    # Entities are flagged for manual review
    if c_Type == "E":
        # entity, flag for manual review
        print(r, "Entity: Review manually", end='\r')
        ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
        ws.cell(row=r, column=head.col['Remarks'], value='Entity')
        print(r, "Entity.                                 ", end='\r')
        continue
    if DebugFlg:
        print(r, c_Reports)
        input('Enter > ')
    if "CRIME - ORGANIZED" in c_categories:
        ws.cell(row=r, column=head.col['Status'], value='SIC TAG CORRECT')
        ws.cell(row=r, column=head.col['Remarks'], value='Category.')
        print(r, "SIC Correct.                         ", end='\r')
        continue        
    if "NONCONVICTION TERROR" in c_categories:
        ws.cell(row=r, column=head.col['Status'], value='TAG SHOULD BE REMOVED')
        ws.cell(row=r, column=head.col['Remarks'], value='Terror Caterogy found.')
        print(r, "SIC Correct.                         ", end='\r')
        continue        
    if "CRIME - TERROR" in c_categories:
        ws.cell(row=r, column=head.col['Status'], value='TAG SHOULD BE REMOVED')
        ws.cell(row=r, column=head.col['Remarks'], value='Terror Caterogy found.')
        print(r, "SIC Correct.                         ", end='\r')
        continue            
    if not c_Reports:
        ws.cell(row=r, column=head.col['Status'], value='SIC TAG NOT FOUND')
        ws.cell(row=r, column=head.col['Remarks'], value='No report column')
        print(r, "No report found.                         ", end='\r')
        continue
    if len(c_Reports) > 800:
        ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
        ws.cell(row=r, column=head.col['Remarks'], value='Report or list content too long')
        print(r, "Report too long.                         ", end='\r')
        continue
    # check if in additional lists
    if c_OfficialLists:
        if DebugFlg:
            print(r, "List found")
        l_list = c_OfficialLists.split(';')
        i=0
        for tag in l_list:
            # look for tag in c_AdditionalInfo and extract string
            # for fraud, if HSS is found, tag a CORRECT HSS and no further processing
            regex = '\['+tag+'\].*?\['
            #_DEBUG print (regex)
            p = re.compile(regex)
            x = p.search(c_AdditionalInfo)
            if x:
                TagStr.append(x.group())
                 # we do not need to strip the brackets
                print(r, "List match ", tag, "found            ", end="\r")
                i += 1
                Extra = True
            # end if
        # end for
    # end if (OfficialLists)
        if DebugFlg:
            print(TagStr)
    # end if (lists)
    # we now have TagInfo populated
    # len(TagInfo) = mumber of elements (>0) or i , Ectra = True as flag, and i as the 
    # for str in TagInfo:
    #  we check crimes and convictions there as well at the end ofr the following loop
    # check for convvicted crimes in Report
    sic_crime = check_issues(crimes, c_Triage, r)
    # now we check for Addidional List Tags
    if sic_crime == True:
        continue # go to next record
    # now ehck in Bio
    c_Triage = c_Bio
    sic_crime = check_issues(crimes, c_Triage, r)
    if sic_crime:
        continue
    # check in lists
    if Extra:
        print("Checking additional in Lists", end="\r")
        LongReport = False
        ListCheck = True
        for x_Triage in TagStr:
            if len(x_Triage) > 720:
                LongReport = True
                ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
                ws.cell(row=r, column=head.col['Remarks'], value="List entry too long.")
                print(r, "Long list enry.                         ", end='\r')
                continue
            sic_crime = check_issues(crimes, x_Triage, r)
            if sic_crime == True:
                break
        # end for
    # end if
    # if sic_crime was true, it was already written
    if sic_crime == True:
        continue
    if sic_crime == False:
        if preconv_option:
            print(r, 'SIC incorrect                                   ', end='\r')
            ws.cell(row=r, column=head.col['Status'], value="TAG SHOULD BE REMOVED")
            ws.cell(row=r, column=head.col['Remarks'], value="No offence tag found")
            continue
        if not LongReport:
            print(r, "SIC incorrect                              ", end='\r')
            if pre_conv == False:
                ws.cell(row=r, column=head.col['Status'], value="TAG SHOULD BE REMOVED")
                ws.cell(row=r, column=head.col['Remnrks'], value="No offence tag found")
            else:
                ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
                ws.cell(row=r, column=head.col['Remarks'], value="Offence found with possible aquittals")
    if sic_crime == None:
        print(r, "Review manually                            ", end='\r')
        ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
        ws.cell(row=r, column=head.col['Remarks'], value="Large distance between offence and conviction")
    
    # end loop through rows
# write to new workbook
print('\nWriting and saving results in file', dest_file, '...' )
try:
    wb.save(dest_file)
except:
    input("\nCannot write to file. Try to close it first and press enter > ")
    print("Saving...")
    wb.save(dest_file)
print('Done')
# end program ######################################################################################################
 