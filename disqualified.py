#######################################################################
# parse Excel Worksheet for correct SIC tag
# LOGIC FOR DISQUALIFIED OR DEBARRED
# (c) 2022 Theys Radmann
# ver 1.0
# ver 1.1 added triage and changed Entities for tag should be removed
#######################################################################
# modules/libararies needed
from openpyxl import load_workbook, Workbook
import re  # regex
import sys
import argparse
from common import ExcelFile, ExcelHeader, RegexSearch
# definition of crime categories
# order of some crimes in list is important for logic and efficiency
# first crime found and convicted for, ends check for further crimes, that's why
# most frequent ones should be first 
ver = '1.1'
program_name = 'disqualified'
# triage (regular expressions)
crimes = [
    r"(banned|barred|prohibited) from (holding|seeking) public (office|employment)",
    r"banned to hold public office",
    r"banned from( public)? office",
    r"ban imposed by",
    r"imposed? a ban",
    r"imposed a lifetime (trading|director|officer|investor relations) ban",
    r"prohibited from (acting|trading|holding contracts with the public authority)",
    r"disqualif(y|ied)",
    r"debarred",
    r"barred from (acting|applying|association|operating|participating|serving)",
    r"(barred|banned) from( any)? securities industry",
    r"barred from holding executive positions",
    r"barred from practi[sz]ing law",
    r"director bar imposed",
    r"ban imposed by",
    r"banned from contracts? with the state",
    r"banned from dealing in securities",
    r"disbarred (as|by|from)",
    r"barred by",
    r"permanently banned (by|from)",
    r"barred from penny stock",
    r"barred from exerci[sz]ing public functions",
    r"expelled from the cpc",
    r"prohibited from offering and selling securities"
]
aquittals = [
    r"ac?quitt(al|ed)",
    r"pardon(ed)?",
    r"case filed",
    r"dismissed",
    r"dropped"
]
dismissals = [
    r"dismiss(ed|al)",
    r"dropped",
    r"case filed"
]

words_apart = 20 # maximum distance of words apart from crime and conviction when matching cirme frst and conviction second
max_rep_length = 800 # maximum report length for processing, longer that this will get tagged for review
pre_conv = False
DebugFlg = False

# functions
####################################################################
def check_list_sic(list_tag, r):
#  returns true or false for trapping positive sic lists
####################################################################
# lists that trigger positive sic tag (could be read from a file)
    TrueLists =[
        r"ACNV",
        r"ADB",
        r"AFCB",
        r"AFNPA",
        r"AIIB",
        r"DISQUALIFIED DIRECTORS",
        r"INMCA-DD",
        r"MXSFP",
        r"NIB",
        r"PDGS",
        r"PECG",
        r"RUFTS-DD",
        r"UGPPDA",
        r"USDTC",
        r"WORLD BANK"
    ]
    list_status = [ False, "Null"]
    for str_list in TrueLists:
        if str_list in list_tag:
            list_status = [ True , str_list]
            return list_status
    return list_status 
# end check_sic_list()
############################################################
def check_conviction(type, str_report, n):
# returning True, False, or None
# checks if there was a convitcion for the crime type
# type : crime (string)
# str_report : record (report column) (string)
# r: row begin processed, for informational purposes only (debugging)
# returns 1 if found and issue follows conviction (writes)
# returns 2 if found and issue is followed by conviction (writes)
# returns -1 if issue is followed by conviction but too far apart (no write)
# returns -2 if conviction found with reversal (acquittal) found (writes)
# returns 0 is no conviction was found at all (no write)
############################################################
    post_conv = 0
    long_flag = False
    global words_apart, DebugFlg, preconv_option
    phrase = [
        r"found guilty",
        r"convicted",
        r"sentence[d]?",
        r"pleaded guilty",
        r"pleaded no contest",
        r"imprisoned",
        r"fined",
        r"arrested .+ serve",
        r"for conviction",
        r"ordered .*\s*to (pay|serve)",
        r"incarcerated",
        r"admitted guilt",
        r"served probation",
        r"to serve .* imprisonment",
        r'previous conviction[s]* .*?'
    ]
    # keywords must be near crime type if before conv
    # build search string with crime type
   
    for str in phrase:
        long_flag = False
        # search crime after conviction. Distance of words are not checked as crime usually follows conviction after a few words
        #  if specified after.
        s_str = str + ' .*?' + type  # RegEx word followed by space and anythnig in between and the second word
        if DebugFlg:
            print(n, s_str)
            input("Press return ")
        x = RegexSearch(s_str, str_report, n)
        if x: 
            # conviction found
            if DebugFlg:
                print(n, x.group())
            words = re.split('\s', x.group())
            if len(words) > words_apart:
                # crime too far from sentence, look for another sentence further ahead, in case there are two
                n_idx = slice(x.start()+1, len(str_report)-1)
                y =RegexSearch(s_str, str_report[n_idx], n)
                if y:
                    words = re.split("\s", y.group())
                    if len(words) > words_apart:
                        print (n, "Too many words between issue and conv", end="\r")
                        post_conv = -1 # to flag for review
                        print(r, "SIC Review                            ", end='\r')
                        ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
                        if ListCheck:
                            ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: Remote connection between conviction and offence. From List")
                        else:
                            ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: Remote connection between conviction and offence")
                        return True                       # issue is too far for a conclusive conviction
                    # issue is too far for a conclusive conviction
                # let's ignore this., but flag for review in case we do not find further evidence (return -1)
            # end if (len)
            # we have foud conviction, check for aquittals
            for tag in aquittals:
                s_str = type + '.*' + tag # may be let's not check for crime type as well
                s_aquitt = RegexSearch(s_str, str_report, n)
                if s_aquitt:
                    print("Dismissal found                                ", end='\r')   
                    return -2
                    # this may be revised
            # end for (aquitals)
            return 2
            
        # if not found, check the other way around. problem is if there was a conviction for somthing different, in which
        # case we should not check for preious mentions of issues
        # this is difficult. Here some tries just to catch these common ones
        if "sentenced for" in str_report:
            continue
        if "pleaded guilty to" in str_report:
            continue
        if "found guilty for" in str_report:
            continue
        if "pleaded no contest to" in str_report:
            continue
        s_str = 'sentenced for \d+ years'
        x = RegexSearch(s_str, str_report, n)
        if not x:
            s_str = "sentence[d]* .*? *for "
            x = RegexSearch(s_str, str_report, n)
            if x:
                continue
        else:
            if DebugFlg:
                print(type, "\n", str)
                print(n, "sentenced for x years found")
                input("enter")
        s_str = "sentence[d]* .*? *on charges of"
        x= RegexSearch(s_str, str_report, n)
        if x:
            continue
        s_str = "found guilty .*? *on charges of"
        x= RegexSearch(s_str, str_report, n)
        if x:
            continue
        s_str = 'pleaded guilty .*? *to'
        x= RegexSearch(s_str, str_report, n)
        if x:
            continue
         # now the other way around, conviction after crime
        s_str = type + r'.*? ' + str
        x = RegexSearch(s_str, str_report, n)
        if x:
            if DebugFlg:
                print(n, x.group())
            # found. if there are too many words between the type and the conviction phrase, assume review
            # there may be a later repetiion of the word wich decreases the word count, do we check twice
            words = re.split("\s", x.group()) # split into words

            if len(words) > words_apart: # too many words in between, but there could be further mention of issue
                # look for issue further ahead
                n_idx = slice(x.start()+1, len(str_report)-1)
                y = RegexSearch(s_str, str_report[n_idx], n)
                if y:
                    words = re.split("\s", y.group())
                    if len(words) > words_apart:
                        print (n, "Too many words between issue and conv", end="\r")
                        post_conv = -1 # to flag for review
                        long_flag = True
                # end if (y)
            # end if (Len)
            if long_flag == False:
                for tag in aquittals:
                    s_str = type + '.*' + tag
                    s_aquitt = RegexSearch(s_str, str_report, n)
                    if s_aquitt:
                        print(n, "Aquittal found                                ", end='\r')   
                        return -2 
                    # this may be revised
                # end for (aquitals)
            # end if (long_flag)
            # found without aquittal
            return 1
        # end if (x)
    # end for (str)
    return post_conv
#####################################################################
def check_item(item, str_Triage, r, pre_conv, src_text):
# normally called by check_issues()
#####################################################################
    global DebugFlg, ListCheck, ws, dismissals
    sic_tag = False
    s_crime = RegexSearch(item, str_Triage, r)
    if s_crime:
        # check for weapons of mass destruction
        if "weapons of mass destruction" in src_text:
            # invalidates crime found
            return False
        if pre_conv:
            # check for dismissed
            for tag in dismissals:
                s_str = item + '.*' + tag # we may ommit item in search string
                s_diss = RegexSearch(s_str, str_Triage, r)
                if s_diss:
                    print(r, "Dismissal found                                ", end='\r')   
                    ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
                    ws.cell(row=r, column=head.col['Remarks'], value="Pre Conv: Dismissal found ["+src_text+"]")
                    xls.review += 1
                    return True
                # end if
            # end for
            print(r, 'SIC correct                                            ', end='\r')
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
            ws.cell(row=r, column=head.col['Remarks'], value="Pre Conv ["+src_text+"]")
            xls.sic_correct += 1
            return True
        # end if preconv
        # post conv processing
        chk = check_conviction(item, str_Triage, r)
        if chk == -1:
            # too far away, flag for review (for now)
            print(r, "SIC Review                     ", end='\r')
            sic_tag = None
        if chk == 1:
            # write correct to sheet
            print(r, "SIC Correct                             ", end='\r')
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
            ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: Conviction after Triage ["+src_text+"]")
            xls.sic_correct += 1
            return True
        if chk == 2:
            # write correct to sheet
            print(r, "SIC Correct                             ", end='\r')
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
            ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: Conviction before Triage ["+src_text+"]")
            xls.sic_correct += 1
            return True
        if chk == -2:
            print(r, 'Review manually                                ', end='\r')
            ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
            ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: acquittal found ["+src_text+"]")
            xls.review += 1
            return True
        # if no conviction was found, check if there was a dismissal
        if chk == 0:
            for tag in dismissals:
                s_str = item + '.*' + tag # we may ommit item in search string
                s_diss = RegexSearch(s_str, str_Triage, r)
                if s_diss:
                    print(r, "Dismissal found                                ", end='\r')   
                    ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
                    ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: No conviction, dismissal found ["+src_text+"]")
                    xls.review += 1
                    return True # behaves like correct as no further offences are affeced if there is a dismissal
                    # this may be revised
                # end if
            # end for
            print (r, 'Tag correct - pre conv', end='\r')
            ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
            ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: No conviction ["+src_text+"]")
            xls.review += 1
            return True

    # end if (s_crime true)
    return sic_tag  # can only be False or None here
        

#####################################################################
def check_issues(issues, str_Triage, r, preconv, Source):
# checks if crime was found and convicted
# # returns True if offense found
# (crime found and written in record in check_item()), 
# False (no tag/offense found), 
# and None (to review, long list or report, or distance) 
# 
#####################################################################
    global DebugFlg, ws
    sic_tag = False
    # first, let's check review tags

    for x_crime in issues:
        sic_tag = check_item(x_crime, str_Triage, r, preconv, Source)
        if sic_tag:
            break
    # end for (issues loop)
    return sic_tag
# end functions
###################################################

# start program
Testflag = False
DebugFlg = False
TrueCondition = False
offence_found = False
row_limit = 0
sic_tag = False

xls = ExcelFile(program_name, ver)

if xls.debug_flag:
    DebugFlg = True
if xls.test_flag:
    row_limit = xls.row_limit
    Testflag = True
    
ws = xls.ws
head = ExcelHeader(ws)

if DebugFlg:
    print(head.col)
    input('Enter > ')
# for each row
r = 0
for row in ws.rows:
    pre_conv = False
    r += 1
    if r == 1:
        continue
    if Testflag:
        if r > row_limit:
            break
    # read row into variables (only useful ones)
    c_categories = ws.cell(row=r, column=head.col['Categories']).value       
    c_OfficialLists = ws.cell(row=r, column=head.col['OfficialLists']).value    
    c_AdditionalInfo = ws.cell(row=r,column=head.col['AdditionalInfo']).value
    c_Reports = ws.cell(row=r,column=head.col['Reports']).value     
    c_Type = ws.cell(row=r, column=head.col['Type']).value
    # c_Bio = ws.cell(row=r, column=head.col['Bio']).value
    c_status= ws.cell(row=r, column=head.col['Status']).value
    c_Triage = c_Reports
    c_lists = [] # resets list of OifficialLists
    ListsTrue = False
    LongReport = False
    ListCheck = False
    sic_tag = False
    if head.missing_col == 'Type':
        c_Type = "N"
    
    # Entities are flagged for manual review
    if c_Type == "E":
        # entity, flag for manual review
        print(r, "Entity: Review manually", end='\r')
        ws.cell(row=r, column=head.col['Status'], value="TAG SHOULD BE REMOVED")
        ws.cell(row=r, column=head.col['Remarks'], value='ENTITY')
        print(r, "Entity.                                 ", end='\r')
        xls.entities +=1
        xls.sic_incorrect += 1
        continue
    if DebugFlg:
        print(r, c_Reports)
        input('Enter > ')
    if not c_Reports:
        ws.cell(row=r, column=head.col['Status'], value='NO REPORT')
        ws.cell(row=r, column=head.col['Remarks'], value='No report column')
        print(r, "No report found.                         ", end='\r')
        xls.no_report +=1
        continue
    if c_OfficialLists and c_OfficialLists != "NULL":
        sic_list = check_list_sic(c_OfficialLists, r)
        if sic_list[0] == True:
            print(r, "Tagged list", end='\r')
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
            ws.cell(row=r, column=head.col['Remarks'], value='OFFICIAL LIST : ' + sic_list[1])
            xls.off_lists += 1
            xls.sic_correct += 1
            continue # no further processing needed
        # Check if there are brackets for lists in AdditionalInfo and populate set
    if len(c_Reports) > max_rep_length:
        ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
        ws.cell(row=r, column=head.col['Remarks'], value='LONG CONTENT')
        print(r, "Report too long.                         ", end='\r')
        xls.long_entries +=1
        xls.review += 1
        continue
    # check if in additional lists
    if c_OfficialLists and c_OfficialLists != "NULL":
        if DebugFlg:
            print(r, "List found")
        l_list = c_OfficialLists.split(';')
        i=0
        for tag in l_list:
            # look for tag in c_AdditionalInfo and extract string
            regex = '\['+tag+'\].*?\['
            #_DEBUG print (regex)
            x = RegexSearch(regex, c_AdditionalInfo, r)
            if x:
                c_lists.append(x.group())
                 # we do not need to strip the brackets
                print(r, "List match ", tag, "found            ", end="\r")
                i += 1
                ListsTrue = True
            # end if
        # end for
    # end if (OfficialLists)
    if DebugFlg:
        print(c_lists)
    
    # we now have TagInfo populated
    # Review keywords
        # flag pre/post conv
    if "CRIME" in c_categories:
        pre_conv = False
        xls.postconv +=1
    else:
        pre_conv = True
        xls.preconv +=1
    sic_tag = check_issues(crimes, c_Triage, r, pre_conv, 'RPT')
    if sic_tag == True:
        continue # go to next record
    # check in lists
    if ListsTrue:
        print("Checking additional in Lists", end="\r")
        ListCheck = True
        for x_Triage in c_lists:
            LongReport = False
            if len(x_Triage) > max_rep_length:
                LongReport = True
                ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
                ws.cell(row=r, column=head.col['Remarks'], value="LONG REPORT [LIST]")
                print(r, "Long list enry.                         ", end='\r')
                xls.long_entries += 1
                xls.review += 1
                continue # next list entry
            if sic_tag:
                break # no more list cheks
            sic_tag = check_issues(crimes, x_Triage, r, pre_conv, 'LIST')
            if sic_tag == True:
                break
        # end for (list)
    # end if (extra/ lists)
    # if sic_tag was true, it was already written
    if sic_tag == True:
        continue
    if sic_tag == False:
        if LongReport:
            print (r, "Review manually.                                 ", end='\r')
            ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
            ws.cell(row=r, column=head.col['Remarks'], value="List Content to long.")
            xls.long_entries += 1
            xls.review += 1
            continue
        if pre_conv:
            print(r, 'SIC incorrect                                   ', end='\r')
            ws.cell(row=r, column=head.col['Status'], value="TAG SHOULD BE REMOVED")
            ws.cell(row=r, column=head.col['Remarks'], value="Pre Conv: No offence found")
            xls.sic_incorrect += 1
            continue
        print(r, 'SIC incorrect                                   ', end='\r')
        ws.cell(row=r, column=head.col['Status'], value="TAG SHOULD BE REMOVED")
        ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: No offence found")
        xls.sic_incorrect += 1
    if sic_tag == None:
        print(r, "Review manually                            ", end='\r')
        ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
        ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: relation between crime and conviction not clear")
        xls.review += 1
    # end loop through rows
# write to new workbook
print('\nWriting and saving results in file', xls.dest_file, '...' )
try:
    xls.ExcelSave()
except:
    input("\nCannot write to file. Try to close it first and press enter > ")
    print("Saving...")
    xls.ExcelSave()
print('Done')
print('Summary')
print('=======')
print('Entities:\t',xls.entities)
print('Long Entries:\t',xls.long_entries)
print('Official Lists:\t', xls.off_lists)
print('No Report:\t',xls.no_report)
print('Pre Conv:\t', xls.preconv)
print('Post Conv:\t',xls.postconv)
print('SIC Correct:\t',xls.sic_correct)
print('SIC Incorrect:\t',xls.sic_incorrect)
print('Man. Review:\t',xls.review)
print('Total:\t\t', r)
# end program ######################################################################################################
 