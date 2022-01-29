#######################################################################
# parse Excel Worksheet for correct SIC flag
# LOGIC FOR FRAUD
# crime must match and must be convicted for it
# Parsing Reports Column (H) and writing results in column 15 ()
# (c) 2022 Theys Radmann
# ver 1.0 commited 202
# Notes: TODO refine logic with client
#######################################################################
# modules/libararies needed
from openpyxl import load_workbook, Workbook
import re  # regex
import sys
import argparse
# definition of crime categories
# order of some crimes in list is important for logic and efficiency
# first crime found and convicted for, ends check for further crimes, that's why
# put most frequent ones first 
ver = '1.0'
crimes = [
    r"fraud",
    r"defraud",
    r"false invoice[s]*",
    r"false claim[s]*",
    r"false declaration[s]*",
    r"false financial",
    r"false health care",
    r"false represenation",
    r"false.*? claim[s]*",
    r"ponzi",
    r"pyramid scheme",
    r"simulated operation",
    r"cheating",
    r"dupe[ei]\w+",
    r"false preten\w+",
    r"patient brokering",
    r"sinulated operation[s]*",
    r"develop false businesses",
    r"dishonest",
    r"kickback",
    r"payday loan collection scheme",
    r"theft by deception",
    r"identity theft",
    r"theft of .*identi\w+",
    r"stolen identi\w+",
    r"violat.* medicare rules",
    r"violation of fraudulent",
    r"sale of unregistered .*without .* *registration"
    ]
words_apart = 20 # maximum distance of words apart from crime and conviction when matching cirme frst and conviction second
pre_conv = False
DebugFlg = False
# functions
############################################################
def check_conviction(type, str_report, n):
# returning True, False, or None
# checks if there was a convitcion for the crime type
# type : crime (string)
# str_report : record (report column) (string)
# r: row begin processed, for informational purposes only (debugging)
# returns 1 if found and issue follows conviction
# returns 2 if found and issue is followed by conviction
# returns -1 if issue is folloed by conviction but too far apart
# returns 0 is no conviction was found at all
############################################################
    post_conv = 0
    long_flag = False
    global words_apart, DebugFlg, preconv_option
    phrase = [
        r"convicted",
        r"sentence[d]*",
        r"pleaded guilty",
        r"pleaded no contest",
        r"found guilty",
        r"imprisoned",
        r"fined",
        r"arrested .+ serve",
        r"for conviction",
        r"ordered .*\s*to (pay|serve)",
        r"incarcerated",
        r"admitted guilt",
        r"served probation",
        r"to serve .* imprisonment",
        r'previous conviction[s]* .*?'
    ]
    # keywords must be near crime type if before conv
    # build search string with crime type
   
    for str in phrase:
        long_flag = False
        # search keyword after conviction. Distance of words are not checked as crime usually follows conviction after a few words
        #  if specified after.
        s_str = str + ' .*?' + type  # RegEx word followed by space and anythnig in between and the second word
        if DebugFlg:
            print(n, s_str)
            input("Press return ")
        p = re.compile(s_str, re.IGNORECASE)
        x = p.search(str_report)
        if x:
            if DebugFlg:
                print(n, x.group())
            if len(x.group()) > 100:
                # issue is too far for a conclusive conviction
                # let's ignore this., but glag for review in ae we do not find further evidence (return -1)
                long_flag = True
            else:
                return 1 # exit fucntion for any crime found. these are most cases.
        # if not found, check the other way around. problem is if there was a conviction for somthing different, in which
        # case we should not check for preious mentions of issues
        # this is difficult. Here some tries just to catch these common ones
        if "sentenced for" in str_report:
            continue
        if "pleaded guilty to" in str_report:
            continue
        if "found guilty for" in str_report:
            continue
        if "pleaded no contest to" in str_report:
            continue
        s_str = 'sentenced for \d+ years'
        p = re.compile(s_str, re.I)
        x = p.search(str_report)
        if not x:
            s_str = "sentence[d]* .*? *for "
            p = re.compile(s_str, re.I)
            x = p.search(str_report)
            if x:
                continue
        else:
            if DebugFlg:
                print(type, "\n", str)
                print(n, "sentenced for x years found")
                input("enter")
        s_str = "sentence[d]* .*? *on charges of"
        p= re.compile(s_str, re.I)
        x= p.search(str_report)
        if x:
            continue
        s_str = "found guilty .*? *on charges of"
        p= re.compile(s_str, re.I)
        x= p.search(str_report)
        if x:
            continue
        s_str = 'pleaded guilty .*? *to'
        p= re.compile(s_str, re.I)
        x= p.search(str_report)
        if x:
            continue
         # now the other way around
        s_str = type + r'.*? ' + str
        p = re.compile(s_str, re.IGNORECASE)
        x = p.search(str_report)
        if x:
            if DebugFlg:
                print(n, x.group())
            # found. if there are too many words between the type and the conviction phrase, assume review
            # there may be a later repetiion of the word wich decreases the word count, do we check twice
            words = re.split("\s", x.group()) # split into words

            if len(words) > words_apart: # too many words in between, but there could be further mention of issue
                # look for issue further ahead
                y = p.search(str_report, x.start())
                if y:
                    words = re.split("\s", y.group())
                    if len(words) > words_apart:
                        print (n, "Too many words between issue and conv", end="\r")
                        post_conv = -1 # to flag for review
                else:
                    if long_flag:
                        return -1
                    else:
                        return 2 # conviction after issue
            else:
                if long_flag:
                    return -1
                else:
                    return 2 
            # end if
        else:
            pass
        # end if
    # end for
    return post_conv
# end function ######################################################
#####################################################################
def check_issue(issues, str_Triage, r):
# checks if crime was found and convicted
# # returns True (crime found and written in record), False, and None (to review) 
# writes record if correct
#####################################################################
    global pre_conv, preconv_option, DebugFlg
    sic_crime = False
    for x_crime in issues:
        p = re.compile(x_crime, re.I) # to ignore case
        s_crime = p.search(str_Triage)
        if s_crime:
            # put exlusions here
            # end excluisions ##########
            # check conviction for crime
            pre_conv = True # crime found, no conviction (yet)
            # return true if pre-convitcion only requested
            if preconv_option:
                print(r, "SIC Correct                             ", end='\r')
                ws.cell(row=r, column=34, value="CORRECT PRE CONV")
                return True
            chk = check_conviction(x_crime, str_Triage, r)
            if chk == -1:
                # too far away, flag for review (for now)
                print(r, "SIC Review                     ", end='\r')
                sic_crime = None
                # we do not break as we could find a valid record for another kewword
            if chk == 1:
                # write correct to sheet
                print(r, "SIC Correct                             ", end='\r')
                ws.cell(row=r, column=34, value="CORRECT CONV")
                return True
            if chk == 2:
                # write correct to sheet
                print(r, "SIC Correct                             ", end='\r')
                ws.cell(row=r, column=34, value="CORRECT INF")
                return True
                    
        # end if
    # end for
    return sic_crime
# end funcfions
###################################################

# start program
DebugFlg = False
preconv_option = False
parser = argparse.ArgumentParser(description='Process Fraud SIC', prog='fraud')
parser.add_argument("--pc", help="Chcek pre-convition only)", action='store_true')
parser.add_argument("--version",help="Displays version only", action='version', version='%(prog)s ' + ver)
parser.add_argument("--debug", help="Debug mode", action='store_true')
parser.add_argument('filename', help="filename to read")
args = parser.parse_args()
if args.debug:
    DebugFlg = True
if args.pc:
    preconv_option = True
    print("Pre conviction option")
org_file = args.filename
if ".xlsx" not in org_file:
    if preconv_option:
        dest_file = org_file + ' Preconv Passed.xlsx'
    else:
        dest_file = org_file + ' Passed.xlsx'
    WorkSheet = org_file
    org_file = org_file + '.xlsx'
else:
    file_parts = org_file.split('.')
    WorkSheet = file_parts[0]
    if preconv_option:
        dest_file = file_parts[0] + ' Preconv Passed.xlxs'
    else:
        dest_file = file_parts[0] + ' Passed.xlxs'
    if DebugFlg:
        print('Org: ', org_file)
        print('Dest: ', dest_file)
        sys.exit()
# open workbook
print( 'Loading spreadsheet ', org_file)
# check if filename exists
#
try:     
    wb = load_workbook(filename=org_file)
except:
    print("cannot open file ", org_file)
    sys.exit()
if preconv_option:
    ws = wb[WorkSheet]
else:
    ws = wb[WorkSheet]
r = 0
print("Processing worksheet")
for row in ws.rows:
    r += 1
    pre_conv = False
    hhs = False
    if r == 1:
        continue # skip header (first row)
    # read row into variables (only useful ones)
    c_categories = ws.cell(row=r, column=7).value       # column G
    c_OfficialLists = ws.cell(row=r, column=8).value    # column H
    c_AdditionalInfo = ws.cell(row=r,column=9).value    # cloumn I
    c_Reports = ws.cell(row=r,column=12).value          # column L
    c_Type = ws.cell(row=r, column=25).value            # column Y
    c_status= ws.cell(row=r, column=34).value           # column AH
    c_Triage = c_Reports
    TagStr = [] # resets list of OifficialLists
    Extra = False
    LongReport = False
    # skip non-crime records
    #if "CRIME" not in c_categories:
    #    continue

    # Note: generalisation: one could filter only for review manaully records (c_status == "REVIEW MANUALLY"). e.g.:
    # if "REVIEW" not in c_status:
    #       continue
    # Entities are flagged for manual review
    if c_Type == "E":
        # entity, flag for manual review
        print(r, "Entity: Review manually", end='\r')
        ws.cell(row=r, column=34, value="ENTITY: REVIEW MANUALLY")
        print(r, "Entity.                                 ", end='\r')
        continue
    if DebugFlg:
        print(r, c_Reports)
        input('Enter > ')
    if not c_Reports:
        ws.cell(row=r, column=34, value='NO REPORT')
        print(r, "No report found.                         ", end='\r')
        continue
    if len(c_Reports) > 720:
        ws.cell(row=r, column=34, value="LONG REPORT: REVIEW MANUALLY")
        print(r, "Report too long.                         ", end='\r')
        continue
    # check if in additional lists
    if c_OfficialLists:
        # extract lists from string
        # split string
        
        if DebugFlg:
            print(r, "List found")
        l_list = c_OfficialLists.split(';')
        i=0
        for tag in l_list:
            # look for tag in c_AdditionalInfo and extract string
            # for fraud, if HSS is found, tag a CORRECT HSS and no further processing
            if "HHS" in tag:
                print(r, "HSS fouund in lists.                     ", end='\r')
                ws.cell(row=r, column=34, value="CORRECT HHS" )
                hhs = True
                break
            regex = '\['+tag+'\].*?\['
            #_DEBUG print (regex)
            p = re.compile(regex)
            x = p.search(c_AdditionalInfo)
            if x:
                TagStr.append(x.group())
                 # we do not need to strip the brackets
                print(r, "LIst match ", tag, "found            ", end="\r")
                i += 1
                Extra = True
            # end if
    #   # end for
        if DebugFlg:
            print(TagStr)
    # end if (lists)
    # we now have TagInfo populated
    # len(TagInfo) = mumber of elements (>0) or i , Ectra = True as flag, and i as the 
    # for str in TagInfo:
    #  we check crimes and convictions there as well at the end ofr the following loop
    if hhs:
        continue # next record
    # check for convvicted crimes in Report
    sic_crime = check_issue(crimes, c_Triage, r)
    # now we check for Addidional List Tags
    if sic_crime == True:
        continue # go to next record
    if Extra:
        print("Checking additional in Lists", end="\r")
        LongReport = False
        for x_Triage in TagStr:
            if len(x_Triage) > 720:
                LongReport = True
                ws.cell(row=r, column=34, value="LONG LIST ENTRY: REVIEW")
                print(r, "Long list enry.                         ", end='\r')
                continue
            sic_crime = check_issue(crimes, x_Triage, r)
            if sic_crime == True:
                break
        # end for
    # end if
    # if sic_crime was true, it was already written
    if sic_crime == True:
        continue
    if sic_crime == False:
        if not LongReport:
            print(r, "SIC incorrect                              ", end='\r')
            if pre_conv == False:
                ws.cell(row=r, column=34, value="INCORRECT")
            else:
                ws.cell(row=r, column=34, value="INCORRECT NO CONV (REVIEW)")
    if sic_crime == None:
        print(r, "Review manually                            ", end='\r')
        ws.cell(row=r, column=34, value="REVIEW MANUALLY")
    
    # end loop through rows
# write to new workbook

print('\nWriting and saving results spreadsheet ', dest_file, '...')
try:
    wb.save(dest_file)
except:
    input("Cannot write to file. Try to close it first and press enter > ")
    print("Saving...")
    wb.save(dest_file)
print('Done')
# end program ######################################################################################################
 