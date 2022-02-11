#######################################################################
# parse Excel Worksheet for correct SIC flag
# LOGIC FOR ABSCONDER OR FUGITIVE
# crime must match and must be convicted for it
# Parsing Reports Column (H) and writing results in column 15 ()
# (c) 2022 Theys Radmann
# ver 1.0 2022-02-01
# ver 1.1 2022-02-01 changed to incorporte ExcelHeader as generic header reader
# run version 2.0 with pre conv only (meant for crt)
#######################################################################
# modules/libararies needed
from openpyxl import load_workbook, Workbook
import re  # regex
import sys
import argparse
from common import ExcelHeader, RegexSearch
# definition of offence categories
# order of some offences in list impacts efficiency
ver = '1.2'
Triage = [
    r"wanted (by|in)",
    r"abscond(ed)?",
    r"abscond(er|ing)",
    r"at large",
    r"escape(d)?"
    r"fled",
    r"wanted[.]",
    r"flee(s)?"
    r"fugitive(s)?",
    r"in absentia",
    r"most wanted",
    r"uamvs",
    r"unkonwn whereabouts",
    r"uzmvd",
    r"vnmps-mw",
    r"whereabouts( remain)? unknown",
    r"on (.+? ){0,2}?wanted list",
    r"top wanted list",
    r"wanted (for|in)"

    ]
# reverse trieage are those words that invalidate a true a&f keyword. it must follow the offence
ReverseTriage = [
    r"imprisoned",
    r"busted",
    r"captured",
    r"(re)?arrested",
    r"custody",
    r"detained",
    r"apprehended",
    r"surrendered",
    r"electronic monitoring",
    r"electronic surveillance measures",
    r"incarceration",
    r"in prison",
    r"bail granted",
    r"remanded in custody",
    r"remanded",
    r"extradited",
    r"deported",
    r"pleaded( not)? guilty",
    r"granted( .+?){1,2} bail", 
    r"no longer (listed|wanted)",
    r"no longer on (an enforcement )?list",
    r"removed from( .+?){0,3}(wanted|fugitives) list", 
    r"under house arrest",
    r"under( special)? surveillance",
    r"sentenced",
    r"convicted"
]
# we also add triage keywords. These are the list names that appear in AdditionalLists that flag as valid SIC. We read thes from a file

words_apart = 20 # maximum distance of words apart from crime and conviction when matching cirme frst and conviction second
pre_conv = False
DebugFlg = False
# functions
####################################################################
def check_list_sic(list_tag, r):
#  returns true or false for trapping positive sic lists
####################################################################
    global lws, TrueList
    # lists that trigger positive sic tag. could be read from a file 
    for str_list in TrueList:
        if str_list in list_tag:
            return True
    return False 
# end check_sic_list()
#####################################################################
def check_issue(issues, str_Triage, r):
# checks if crime was found and convicted
# # returns True (crime found and written in record), False, and None (to review) 
# writes record if correct
# called from main program, once with report, and one or more times with list entry passed in str_Triage
#####################################################################
    global pre_conv, DebugFlg, ListCheck, ReverseTriage, ReverseTag
    sic_crime = False
    ReverseTag = False
    for x_crime in issues:
        s_crime = RegexSearch(x_crime, str_Triage, r)
        if s_crime:
            sic_crime = True
            # issue found, check for reverse triage kword if it follows
            # there is normally only one issue per record
            for kword in ReverseTriage:
                s_str = x_crime + '.* ' + kword 
 
                s_counter = RegexSearch(s_str, str_Triage, r)
                if s_counter:
                # counter found, exit incorrect
                    ReverseTag = True
                    break
        # end if
    # end for
    if sic_crime:
        if ReverseTag:
            print(r, 'Reivew Manually                             ', end='\r')
            ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
            ws.cell(row=r, column=head.col['Remarks'], value="REVERSE TAG PRESENT")
            return True
        print(r, 'SIC tag correct                             ', end='\r')
        if ListCheck:
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
            ws.cell(row=r, column=head.col['Remarks'], value="From list")
        else:
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
            ws.cell(row=r, column=head.col['Remarks'], value="From report")
        return True
    # end if sic_crime
    return sic_crime
# end functions
###################################################

# start program
DebugFlg = False
preconv_option = False
CheckTag = False
ReverseTag = False
TrueList = []
parser = argparse.ArgumentParser(description='Process Absconder & Fugitive SIC', prog='absconder')
parser.add_argument("--version",help="Displays version only", action='version', version='%(prog)s ' + ver)
parser.add_argument("--debug", help="Debug mode", action='store_true')
parser.add_argument('filename', help="filename to read")
args = parser.parse_args()
if args.debug:
    DebugFlg = True
org_file = args.filename
if ".xlsx" not in org_file:
    dest_file = org_file + ' Passed.xlsx'
    WorkSheet = org_file 
    org_file = org_file + '.xlsx'
else:
    file_parts = org_file.split('.')
    dest_file = file_parts[0] + ' Passed.xlxs'
    WorkSheet = file_parts[0]
if DebugFlg:
    print('Org: ', org_file)
    print('Dest: ', dest_file)
    input('enter > ')
# open workbook

print( 'Loading spreadsheet', org_file)
# check if filename exists
#
try:     
    wb = load_workbook(filename=org_file)
except:
    print("cannot open file", org_file)
    sys.exit()
ws = wb[WorkSheet]
# load lists that enforce offence (from triage)
try:
    lwb = load_workbook(filename='SIC Absconder&Fugitive Logic.xlsx')
except:
    print("cannot open file 'SIC Absconder&Fugitive Logic.xlsx'. Maybe it's open?")
    input('Enter to coninue >')
    lwb = load_workbook(filename='SIC Absconder&Fugitive Logic.xlsx')    
lws = lwb['TRIAGE KEYWORDS']
print('Loading Triage Lists')
r = 0
print(lws.rows)
for lrows in lws.rows:
    r +=1
    if r == 1: 
        continue # skip header
    TrueList.append(lws.cell(row=r, column=1).value)       
r = 0
print("Processing worksheet")
head = ExcelHeader(ws)
for row in ws.rows:
    pre_conv = False
    r += 1
    if r == 1:
        # ws.cell(row=1, column=23, value='Status')
        continue
    # read row into variables (only useful ones)
    c_categories = ws.cell(row=r, column=head.col['Categories']).value       
    c_AdditionalInfo = ws.cell(row=r,column=head.col['AdditionalInfo']).value 
    c_OfficialLists = ws.cell(row=r, column=head.col['OfficialLists']).value   
    c_Reports = ws.cell(row=r,column=head.col['Reports']).value         
    c_Type = ws.cell(row=r, column=head.col['Type']).value           
    c_status= ws.cell(row=r, column=head.col['Status']).value          
    c_Triage = c_Reports
    TagStr = [] # resets list of OifficialLists
    Extra = False
    LongReport = False
    ListCheck = False # presence of list bracket in additionalinfo
    ListSic = False # Triage Keywords list found flag
    # if "CRIME" not in c_categories:
    #    continue

    # Note: generalisation: one could filter only for review manaully records (c_status == "REVIEW MANUALLY"). e.g.:
    # if "REVIEW" not in c_status:
    #       continue
    # Entities are flagged for manual review
    if c_Type == "E":
        # entity, flag for manual review
        print(r, "Entity: Review manually", end='\r')
        ws.cell(row=r, column=head.col['Status'], value="TAG SHOULD BE REMOVED")
        ws.cell(row=r, column=head.col['Remarks'], value="ENTITY")
        print(r, "Entity.                                 ", end='\r')
        continue
    if DebugFlg:
        print(r, c_Reports)
        input('Enter > ')
    if not c_Reports:
        ws.cell(row=r, column=head.col['Status'], value='NO REPORT')
        print(r, "No report found.                         ", end='\r')
        continue
    if DebugFlg:
        print(r, c_Reports)
        input('Enter > ')
    if len(c_Reports) > 800:
        if not CheckTag:
            ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
            ws.cell(row=r, column=head.col['Remarks'], value="LONG RFEPORT")
            print(r, "Report too long.                         ", end='\r')
            continue
    # check if in additional lists
    if c_OfficialLists:
        # check for postive trigger lists
        if (check_list_sic(c_OfficialLists, r)):
            print(r, "Tagged list", end='\r')
            ListSic = True
        # extract lists from string
        # split string
        if DebugFlg:
            print(r, "List found")
        l_list = c_OfficialLists.split(';')
        # store content of official lists tags in AdditionalInfo if there are matches
        i=0
        for tag in l_list:
            # check if list is part of a positive trigger
            # look for tag in c_AdditionalInfo and extract string
            regex = '\['+tag+'\].*?\['
            #_DEBUG print (regex)
            x = RegexSearch(regex, c_AdditionalInfo, r)
            if x:
                TagStr.append(x.group())
                 # we do not need to strip the brackets
                print(r, "List match ", tag, "found            ", end="\r")
                i += 1
                Extra = True
            # end if
    #   # end for
    # end if (lists)
    # we now have everyting polulated
    # 
    if ListSic:
        # positive tag, only need to check for reverse triage
        for str_rev in ReverseTriage:
            m_rev = RegexSearch(str_rev, c_Triage, r)
            if m_rev:
                print(r, "Reverse triage found", end='\r')
                break
        # end for    
        if m_rev:
            ws.cell(row=r, column=head.col['Status'], value='REVIEW MANUALLY')
            ws.cell(row=r, column=head.col['Remarks'], value='OFFICIAL LIST WITH REVERSE')
        else:
            ws.cell(row=r, column=head.col['Status'], value='SIC TAG CORRECT')
            ws.cell(row=r, column=head.col['Remarks'], value='OFFICIAL LIST PRESENT')
        continue
    # end if
    # normal processing
    sic_crime = check_issue(Triage, c_Triage, r)
    if sic_crime == True:
        continue # go to next record
    # now we check for Addidional List Tags
    if Extra:
        print("Checking additional in Lists", end="\r")
        LongReport = False
        ListCheck = True
        for x_Triage in TagStr:
            if len(x_Triage) > 720 and not CheckTag:
                LongReport = True
                ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
                ws.cell(row=r, column=head.col['Remarks'], value="LONG LIST ENTRY")
                print(r, "Long list enry.                         ", end='\r')
                continue
            sic_crime = check_issue(Triage, x_Triage, r)
            if sic_crime == True:
                break
        # end for
    # end if
    # if sic_crime was true, it was already written
    if sic_crime == True:
        continue # to next record
    if sic_crime == False:
        print(r, 'SIC incorrect                                   ', end='\r')
        ws.cell(row=r, column=head.col['Status'], value="TAG SHOULD BE REMOVED")
        ws.cell(row=r, column=head.col['Remarks'], value='NO SIC KEYWORD')
        continue
    if sic_crime == None:
        print(r, "Review manually                            ", end='\r')
        ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
        ws.cell(row=r, column=head.col['Remarks'], value="Remote connection between absconder and recapture")
    
    # end loop through rows
# write to new workbook
# delete the other sheet
print('\nWriting and saving results in file', dest_file, '...' )
try:
    wb.save(dest_file)
except:
    input("\nCannot write to file. Try to close it first and press enter > ")
    print("Saving...")
    wb.save(dest_file)
print('Done')
# end program ######################################################################################################
