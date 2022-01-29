#######################################################################
# parse Excel Worksheet for correct SIC flag
# lOGIC FOR CRIMES
# crime must match and must be convicted for it
# Parsing Reports Column (H) and writing results in column 15 ()
# (c) 2022 Theys Radmann
# ver 1.4 commited 2022-01-16
# Notes: TODO refine logic with client
#######################################################################
# modules/libararies needed
from openpyxl import load_workbook, Workbook
import re  # regex
import sys
import argparse
# definition of crime categories
# order of some crimes in list is important for logic and efficiency
# first crime found and convicted for, ends check for further crimes, that's why
# put most frequent ones first 
ver = '1.'
crimes = [
    r"stolen",
    r"steal",
    r"stole public",
    r"misappropriat\w+",
    r"embezzl\w.*",
    r"peculat\w*",
    r"larceny",
    r"to rob",
    r"robbing",
    r"robber\w*",
    r"grand theft",
    r"theft",
    r"theft of", # must be after theft
    # r"organized crime",
    r"burglar\w*",
    r"pilfering",
    r"heist",
    r"shoplift\w*",
    r"siphon\w*",
    r"(diverted|diversion)", 
    # r"illicit enrichment",
    r"malversat\w+",
    # r"public deposit",
    # r"illegally receiving public",
    r"absorbing public (deposits|funds)",
    r"illegal absorption of public",
    r"illegally (receiving|acquiring) .*deposits",
    r"illegally (receiving|acquiring) .*public",
    r"illegally (accepting|absorbing) .*public",
    r"illegally (accepting|absorbing) .*deposits",
    r"illicit enrichment",
    r"illegally obtain\w*",
    r"illegally receiving pension",
    r"illegally receiving survivor benefits",
    # r"illegal gain",
    r"funds .*\s*illegally",
    # r"financial management irregularities",
    r"larceny",
    r"ditribuit.* state .*\s*asssets"
    # "Misappropriation of funds", included in misappropation
    # "Trafficking in Stolen Goods",
    # "Selling stolen goods",
    # "Receipt of stolen goods",
    # "Intellectual Property",
    ]

words_apart =20 
pre_conv = False
DebugFlg = False
# functions
############################################################
def check_conviction(type, str_report, n):
# returning True, False, or None
# checks if there was a convitcion for the crime type
# type : crime (string)
# str_report : record (report column) (string)
# r: row begin processed, for informational purposes only (debugging)
############################################################
    global words_apart, DebugFlg
    post_conv = 0
    phrase = [
        r"convicted",
        r"sentence[d]*",
        r"pleaded guilty",
        r"found guilty",
        r"pleaded no contest",
        r"imprisoned",
        r"fined",
        r"arrested .+ serve",
        r"ordered .*\s*to (pay|serve)",
        r"incarcerated",
        r"amditted guilt",
        r"served probation",
        r"to serve .* imprisonment"
    ]
    # keywords must be near crime type if before conv
    # build search string with crime type
   
    for str in phrase:
        # search keyword after conviction. Distance of words are not checked as crime usually follows conviction after a few words
        #  if specified after.
        s_str = str + ' .*?' + type # RegEx word followed by space and anythnig in between and the second word
        #_DEBUG print(n, s_str)
        #_DEBUG input("Press return")
        p = re.compile(s_str, re.IGNORECASE)
        x = p.search(str_report)
        if x:
            if DebugFlg:
                print(n, x.group())
            return 1 # exit function for any crime found. these are most cases.
        # if not found, check the other way around. problem is if there was a conviction for something different, in which
        # case we should not check for preious mentions of issues
        # this is difficult. Here some tries just to catch these common ones
        if "sentenced for" in str_report:
            continue
        if "pleaded guilty to" in str_report:
            continue
        if "pleaded no contest to" in str_report:
            continue
        if "found guilty for" in str_report:
            continue
        s_str = "sentence[d]* .*?for "
        p = re.compile(s_str, re.I)
        x = p.search(str_report)
        if x:
            continue
        if "sentenced for" in str_report:
            continue
        if "pleaded guilty to" in str_report:
            continue
        if "found guilty of" in str_report:
            continue
        if "pleaded no contest to" in str_report:
            continue
        s_str = "sentence[d]* .*? *for "
        p = re.compile(s_str, re.I)
        x = p.search(str_report)
        if x:
            continue
        s_str = "sentence[d]* .*? *on charges of"
        p= re.compile(s_str, re.I)
        x= p.search(str_report)
        if x:
            continue
        s_str = "found guilty .*? *on charges of"
        p= re.compile(s_str, re.I)
        x= p.search(str_report)
        if x:
            continue
        s_str = 'pleaded guilty .*? *to'
        p= re.compile(s_str, re.I)
        x= p.search(str_report)
        if x:
            continue
        # now the other way around
        s_str = type + '.*? ' + str
        p = re.compile(s_str, re.IGNORECASE)  
        x = p.search(str_report)
        if x:
            if DebugFlg:
                print(n, x.group())
            # found. if there are too many words between the type and the conviction phrase, assume review
            i_str = re.split(r"\s", x.group()) # split into words
            if len(i_str) > words_apart:
                print (n, "Too many words between crime and conv", end="\r")
                y = p.search(str_report, x.start())
                if y:
                    words = re.split("\s", y.group())
                    if len(words) > words_apart:
                        print (n, "Too many words between issue and conv", end="\r")
                        post_conv = -1 # to flag for review
                else:
                    return 2
            else:
                return 2
            # end if
        else:
           pass
        # end if
    # end for
    return post_conv
# end function ######################################################
#####################################################################
def check_issue(issues, str_Triage, r):
# checks if crime was found and convicted
# # returns True (crime found and written in record), False, and None (to review) 
#
#####################################################################
    global pre_conv, DebugFlg, preconv_option
    sic_crime = False
    for x_crime in issues:
        p = re.compile(x_crime, re.I) # to ignore case
        s_crime = p.search(str_Triage)
        if s_crime:
            # found issue in string
            # exlusions
            if x_crime == "theft":
                x = re.search("[Ii]dentity theft", str_Triage)
                if x:
                    # idenfity theft found. look for another theft not preceded by Identity
                    y = re.search(r"(?<![Ii]dentity) theft", str_Triage)
                    if y == None:
                        # no theft not preceeded by identity was found, i.e. only identity theft was found
                        continue
                # check for theft by deception
                x = re.search(r"[Tt]heft by deception", str_Triage)
                if x:
                    # see if ther are no other thefts apart form deception
                    y = re.search(r"[Tt]heft (?!by deception)", str_Triage)
                    if y == None:
                        # no theft not followed by 'by deecption' was found, i.e. only theft by deception was found
                        continue
                # catch exxlusive 'theft of', otherwise this iteration could be flagged as correct
                x = re.search(r"[Tt]heft of", str_Triage)
                if x:
                    y = re.search(r"[Th]heft (?!of)", str_Triage)
                    if y == None:
                        continue # if found, will be processed in next iteration
            if x_crime == "theft of":
                # exlude theft of identity
                x = re.search("theft of identi", str_Triage)
                if x:
                    # look for negative
                    y = re.search(r"theft of (?!identi)", str_Triage)
                    if y == None:
                        continue # only theft of identity was found
            if x_crime == "stolen":
                # ecxclude specific stolen stuff
                x = re.search(r"stolen identi" , str_Triage)
                if x:
                    # look for negative
                    y = re.search(r"stolen (?!identi)", str_Triage)
                    if y == None:
                        continue # only stolen identity ewas found, no more stolen, continue for next phrase
            if x_crime == "stolen":
                x = re.search("stolen credit card scheme" , str_Triage)
                if x:
                    # look for negative
                    y = re.search(r"stolen (?!credit card scheme)", str_Triage)
                    if y == None:
                        continue
            if x_crime == "stolen":
                x = re.search(r"stolen personal identification" , str_Triage)
                if x:
                    # look for negative
                    y = re.search(r"stolen (?!personal identification)", str_Triage)
                    if y == None:
                        continue
            if x_crime == "steal":
                x = re.search("steal trade secrets" , str_Triage)
                if x:
                    # look for negative
                    y = re.search(r"steal (?!trade secrets)", str_Triage)
                    if y == None:
                        continue
            
            # end exclusions
            # check conviction for crime
            # let]s flag that at least we found the crime
            pre_conv = True
            if preconv_option:
                print(r, "SIC Correct                    ", end='\r')
                ws.cell(row=r, column=15, value="CORRECT PRE CONV")
                return True
            chk = check_conviction(x_crime, str_Triage, r)
            if chk == -1:
                # too far away, flag for review (for now)
                print(r, "SIC Review              ", end='\r')
                sic_crime = None
                # we do not break as we could find a valid record for another kewword
            if chk == 1:
                # write correct to sheet
                print(r, "SIC Correct                    ", end='\r')
                ws.cell(row=r, column=15, value="CORRECT CONV")
                return True
            if chk == 2:
                print(r, "SIC Correct                    ", end='\r')
                ws.cell(row=r, column=15, value="CORRECT INF")
                return True
            #  if chk is zero, is was not found and we continue
        # end if
    # end for
    return sic_crime

# end functions
###################################################
preconv_option = False
DebugFlg = False
# start program
parser = argparse.ArgumentParser(description='Process Fraud & Uttering SIC', prog='fraud')
parser.add_argument("--pc", help="Chcek pre-convition only)", action='store_true')
parser.add_argument("--version",help="Displays version only", action='version', version='%(prog)s ' + ver)
parser.add_argument("--debug", help="Debug mode", action='store_true')
args = parser.parse_args()
if args.debug:
    DebugFlg = True
if args.pc:
    preconv_option = True
# open workbook
print( 'Loading spreadsheet TE2 RECORDS done.xlsx...')    
wb = load_workbook(filename="TE2 RECORDS done.xlsx")
ws = wb['Theft & Embezz Post Conv']
r = 0
print("Processing sheet")
for row in ws.rows:
    r += 1
    pre_conv = False
    if r == 1:
        continue # skip header (first row)
    # read row into variables (only useful ones)
    c_categories = ws.cell(row=r, column=5).value
    c_OfficialLists = ws.cell(row=r, column=6).value
    #_DEBUG print(r, c_OfficialLists)
    c_AdditionalInfo = ws.cell(row=r,column=7).value
    c_Reports = ws.cell(row=r,column=8).value
    c_Type = ws.cell(row=r, column=9).value
    c_status= ws.cell(row=r, column=15).value
    c_Triage = c_Reports
    TagStr = [] # resets list of OifficialLists
    Extra = False
    # skip non-crime records
    # if "CRIME" not in c_categories:
    #     continue

    # Note: generalisation: one could filter only for review manaully records (c_status == "REVIEW MANUALLY"). e.g.:
    # if "REVIEW" not in c_status:
    #       continue
    # Entities are flagged for manual review
    if c_Type == "E":
        # entity, flag for manual review
        print(r, "Entity: Review manually                 ", end='\r')
        ws.cell(row=r, column=15, value="REVIEW MANUALLY (ENTITY)")
        continue
    # check if in additional lists
    if c_OfficialLists:
        # extract lists from string
        # split string
        l_list = c_OfficialLists.split(';')
        i=0
        for tag in l_list:
            # look for tag in c_AdditionalInfo and extract string
            regex = '\['+tag+'\].*?\['
            if DebugFlg:
                print (regex)
            p = re.compile(regex)
            x = p.search(c_AdditionalInfo)
            if x:
                TagStr.append(x.group())
                # we do not need to strip the brackets
                i += 1
                Extra = True
            # end if
        # end for
    # end if
    # we now have TagInfo populated
    # len(TagInfo) = mumber of elements (>0) or i , Extra = True as flag, and i as the number of element
       
    # check for convicted crimes in Report
    sic_crime = check_issue(crimes, c_Triage, r)
    # now we check for Addidional List Tags
    if sic_crime == True:
        continue # go to next record (check_isssue() writes correct SIC)
    if Extra:
        print("Checking additional in Lists", end="\r")
        for x_Triage in TagStr:
            sic_crime = check_issue(crimes, x_Triage, r)
            if sic_crime == True:
                break
        # end for
    # end if
    # if sic_crime was true, it was already written
    if sic_crime == True:
        continue # not really necessary but in case more code comes below in loop...
    if sic_crime == False:
        print(r, "SIC incorrect                         ", end='\r')
        if pre_conv == False:
            ws.cell(row=r, column=15, value="INCORRECT")
        else:
            ws.cell(row=r, column=15, value="NO CONV (REVIEW MANUALLY)")
    if sic_crime == None:
        print(r, "Review manually                       ", end='\r')
        ws.cell(row=r, column=15, value="REVIEW MANUALLY")
   
    # end loop through rows
# write to new workbook

print('\nWriting and saving results spreadsheet Theft Passed.xlsx ...')
if preconv_option:
    wb.save('Theft Preconv Passed.xlsx')
else:
    wb.save('Theft Passed.xlsx')
print('Done')
# end python program ######################################################################################################
 