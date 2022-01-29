#######################################################################
# parse Excel Worksheet for correct SIC flag
# LOGIC FOR NARCOTICS
# crime must match and must be convicted for it
# Parsing Reports Column (H) and writing results in column 15 ()
# (c) 2022 Theys Radmann
# ver 1.0 commited 2022
# Notes: TODO refine logic with client
#######################################################################
# modules/libararies needed
from openpyxl import load_workbook, Workbook
import re  # regex
import sys
import argparse
# definition of crime categories
# order of some crimes in list is important for logic and efficiency
# first crime found and convicted for, ends check for further crimes, that's why
# put most frequent ones first 
ver = '1.0'
crimes = [
    r"(traffick|distribut|import|transport|smuggl).*?(narcotics|drugs)",
    r"(narcotics|(drug[s]?)) (traffick|distribution|import|transport|smuggling)",
    r"(deliver|distribut|cultivat|manufactur).*?((drug[s]?)|(narcotic[s]?))",
    r"((drug[s]?)|(narcotic[s]?)) (deliver|distribut|cultivat|manufactur|supply)",
    r"(sell|suppl).*?(narcotics|drugs)",
    r"dealing in (narcotics|drugs)",
    r"sale of (drugs|narcotics)",
    r"selling (drugs|narcotics)",
    r"(narcotics|(drug[s]?)) (production|precursors|cultivation)",
    r"((drug[s]?)|narcotics) (manufactur|conspiracy|dealing)",
    r"((drug[s]?)|narcotics)( related)? ((offence[s]?)|(crime[s]?|)(charge[s]?))",
    r"produc(e|ing) (drugs|narcotics)",
    r"((drug[s]?)|narcotics) posession.*?(deal|sell|sale|dstribut|traffick|cultivat|suppl)",
    r"posession of (drugs|narcotics).*?(deal|sale|sell|distribut|suppl|traffick|cultivat)",
    r"(distribut|sell).*?((controlled substance[s]?)|cocaine)",
    r"posession for the purpose of trafficking",
    r"narcotics ((charge[s]?)|racket)",
    r"racketeering involving (drugs|narcotics)",
    r"drug conspiracy",
    r"involved in narcotics business",
    r"link between narcotic cartels",
    r"unlawful sale and promotion of prescription drugs",
    r"(smuggl|distribut|sell).*?(ketamine)"
    ]
words_apart = 20 # maximum distance of words apart from crime and conviction when matching cirme frst and conviction second
pre_conv = False
DebugFlg = False
# functions
############################################################
def check_conviction(type, str_report, n):
# returning True, False, or None
# checks if there was a convitcion for the crime type
# type : crime (string)
# str_report : record (report column) (string)
# r: row begin processed, for informational purposes only (debugging)
# returns 1 if found and issue follows conviction
# returns 2 if found and issue is followed by conviction
# returns -1 if issue is folloed by conviction but too far apart
# returns 0 is no conviction was found at all
############################################################
    post_conv = 0
    long_flag = False
    global words_apart, DebugFlg, preconv_option
    phrase = [
        r"found guilty",
        r"convicted",
        r"sentence[d]*",
        r"pleaded guilty",
        r"pleaded no contest",
        r"imprisoned",
        r"fined",
        r"arrested .+ serve",
        r"for conviction",
        r"ordered .*\s*to (pay|serve)",
        r"incarcerated",
        r"admitted guilt",
        r"served probation",
        r"to serve .* imprisonment",
        r'previous conviction[s]* .*?'
    ]
    # keywords must be near crime type if before conv
    # build search string with crime type
   
    for str in phrase:
        long_flag = False
        # search keyword after conviction. Distance of words are not checked as crime usually follows conviction after a few words
        #  if specified after.
        s_str = str + ' .*?' + type  # RegEx word followed by space and anythnig in between and the second word
        if DebugFlg:
            print(n, s_str)
            input("Press return ")
        p = re.compile(s_str, re.IGNORECASE)
        x = p.search(str_report)
        if x:
            if DebugFlg:
                print(n, x.group())
            words = re.split('\s', x.group())
            if len(words) > words_apart:
                # crime too far from sentence, look for another sentence further ahead, in case there are two
                y =p.search(str_report, x.start()+1)
                if y:
                    words = re.split("\s", y.group())
                    if len(words) > words_apart:
                        print (n, "Too many words between issue and conv", end="\r")
                        post_conv = -1 # to flag for review
                        # issue is too far for a conclusive conviction
                        # let's ignore this., but flag for review in case we do not find further evidence (return -1)
                    else:
                        return 1
                long_flag = True
                # issue is too far for a conclusive conviction
                # let's ignore this., but flag for review in case we do not find further evidence (return -1)
            else:
                return 1 # exit fucntion for any crime found. these are most cases.
        # if not found, check the other way around. problem is if there was a conviction for somthing different, in which
        # case we should not check for preious mentions of issues
        # this is difficult. Here some tries just to catch these common ones
        if "sentenced for" in str_report:
            continue
        if "pleaded guilty to" in str_report:
            continue
        if "found guilty for" in str_report:
            continue
        if "pleaded no contest to" in str_report:
            continue
        s_str = 'sentenced for \d+ years'
        p = re.compile(s_str, re.I)
        x = p.search(str_report)
        if not x:
            s_str = "sentence[d]* .*? *for "
            p = re.compile(s_str, re.I)
            x = p.search(str_report)
            if x:
                continue
        else:
            if DebugFlg:
                print(type, "\n", str)
                print(n, "sentenced for x years found")
                input("enter")
        s_str = "sentence[d]* .*? *on charges of"
        p= re.compile(s_str, re.I)
        x= p.search(str_report)
        if x:
            continue
        s_str = "found guilty .*? *on charges of"
        p= re.compile(s_str, re.I)
        x= p.search(str_report)
        if x:
            continue
        s_str = 'pleaded guilty .*? *to'
        p= re.compile(s_str, re.I)
        x= p.search(str_report)
        if x:
            continue
         # now the other way around
        s_str = type + r'.*? ' + str
        p = re.compile(s_str, re.IGNORECASE)
        x = p.search(str_report)
        if x:
            if DebugFlg:
                print(n, x.group())
            # found. if there are too many words between the type and the conviction phrase, assume review
            # there may be a later repetiion of the word wich decreases the word count, do we check twice
            words = re.split("\s", x.group()) # split into words

            if len(words) > words_apart: # too many words in between, but there could be further mention of issue
                # look for issue further ahead
                y = p.search(str_report, x.start()+1)
                if y:
                    words = re.split("\s", y.group())
                    if len(words) > words_apart:
                        print (n, "Too many words between issue and conv", end="\r")
                        post_conv = -1 # to flag for review
                else:
                    if long_flag:
                        return -1
                    else:
                        return 2 # conviction after issue
            else:
                if long_flag:
                    return -1
                else:
                    return 2 
            # end if
        else:
            pass
        # end if
    # end for
    return post_conv
# end function ######################################################
#####################################################################
def check_issue(issues, str_Triage, r):
# checks if crime was found and convicted
# # returns True (crime found and written in record), False, and None (to review) 
# writes record if correct
#####################################################################
    global pre_conv, preconv_option, DebugFlg, ListCheck, CheckTag
    sic_crime = False

    for x_crime in issues:
        p = re.compile(x_crime, re.I) # to ignore case
        s_crime = p.search(str_Triage)
        if s_crime:
            # begin exclusions #########
            # exclude if distance between actions and narcotics/drugs is to long, i.e.
            # the match string is very long
            if len(s_crime.group()) > 100:
                # should mark as review, as we found a remote connection
                sic_crime = None
                continue
            # end exclusions ##########
            # check conviction for crime
            pre_conv = True # crime found, no conviction (yet)
            # return true if pre-conviction only requested
            if preconv_option:
                print(r, "SIC Correct                                ", end='\r')
                if ListCheck:
                    ws.cell(row=r, column=35, value="CORRECT PRE CONV (LIST)")
                else:
                    ws.cell(row=r, column=35, value="CORRECT PRE CONV")
                return True
            chk = check_conviction(x_crime, str_Triage, r)
            if chk == -1:
                # too far away, flag for review (for now)
                print(r, "SIC Review                     ", end='\r')
                sic_crime = None
                # we do not break as we could find a valid record for another kewword
            if chk == 1:
                # write correct to sheet
                print(r, "SIC Correct                             ", end='\r')
                if ListCheck:
                    ws.cell(row=r, column=35, value="CORRECT CONV (LIST)")
                else:
                    ws.cell(row=r, column=35, value="CORRECT CONV")
                return True
            if chk == 2:
                # write correct to sheet
                print(r, "SIC Correct                             ", end='\r')
                if ListCheck:
                    ws.cell(row=r, column=35, value="CORRECT (CONV INFERRRED) (LIST)")
                else:
                    ws.cell(row=r, column=35, value="CORRECT (CONV INFERRRED)")
                return True
                    
        # end if
    # end for
    return sic_crime
# end functions
###################################################

# start program
DebugFlg = False
preconv_option = False
CheckTag = False
parser = argparse.ArgumentParser(description='Process Narcitics` SIC', prog='nacrotics')
parser.add_argument("--pc", help="Chcek pre-convition only)", action='store_true')
parser.add_argument("--version",help="Displays version only", action='version', version='%(prog)s ' + ver)
parser.add_argument("--debug", help="Debug mode", action='store_true')
parser.add_argument("--rt", help="Chhek tag removal", action='store_true')
parser.add_argument('filename', help="filename to read")
args = parser.parse_args()
if args.debug:
    DebugFlg = True
if args.pc:
    preconv_option = True
    print("Pre conviction option")
if args.rt:
    CheckTag = True
    TagFile = ' CRT'
else:
    TagFile = ''
org_file = args.filename
if ".xlsx" not in org_file:
    if preconv_option:
        dest_file = org_file + TagFile + ' Preconv Passed.xlsx'
    else:
        dest_file = org_file + TagFile + ' Passed.xlsx'
    org_file = org_file + '.xlsx'
else:
    file_parts = org_file.split('.')
    if preconv_option:
        dest_file = file_parts[0] + TagFile + ' Preconv Passed.xlxs'
    else:
        dest_file = file_parts[0] + TagFile + ' Passed.xlxs'
if DebugFlg:
    print('Org: ', org_file)
    print('Dest: ', dest_file)
    input('enter > ')
# open workbook
if CheckTag:
    WorkSheet = 'Tag should be removed'
    print('Checking tags to be removed option')
else:
    WorkSheet = 'Rev Manually blank no narc tria'

print( 'Loading spreadsheet', org_file)
# check if filename exists
#
try:     
    wb = load_workbook(filename=org_file)
except:
    print("cannot open file", org_file)
    sys.exit()
ws = wb[WorkSheet]
r = 0
print("Processing worksheet")
for row in ws.rows:
    pre_conv = False
    r += 1
    if r == 1:
        # ws.cell(row=1, column=23, value='Status')
        continue
    # read row into variables (only useful ones)
    c_categories = ws.cell(row=r, column=7).value       # column G
    c_OfficialLists = ws.cell(row=r, column=8).value    # column H
    c_AdditionalInfo = ws.cell(row=r,column=9).value    # cloumn I
    c_Reports = ws.cell(row=r,column=12).value           # column L
    c_Type = ws.cell(row=r, column=25).value            # column Y
    c_status= ws.cell(row=r, column=35).value           # column AI
    c_Triage = c_Reports
    TagStr = [] # resets list of OifficialLists
    Extra = False
    LongReport = False
    ListCheck = False
    # if "CRIME" not in c_categories:
    #    continue

    # Note: generalisation: one could filter only for review manaully records (c_status == "REVIEW MANUALLY"). e.g.:
    # if "REVIEW" not in c_status:
    #       continue
    # Entities are flagged for manual review
    if c_Type == "E":
        # entity, flag for manual review
        print(r, "Entity: Review manually", end='\r')
        ws.cell(row=r, column=35, value="ENTITY: REVIEW MANUALLY")
        print(r, "Entity.                                 ", end='\r')
        continue
    if DebugFlg:
        print(r, c_Reports)
        input('Enter > ')
    if not c_Reports:
        ws.cell(row=r, column=35, value='NO REPORT')
        print(r, "No report found.                         ", end='\r')
        continue
    if len(c_Reports) > 720:
        if not CheckTag:
            ws.cell(row=r, column=35, value="LONG REPORT: REVIEW MANUALLY")
            print(r, "Report too long.                         ", end='\r')
            continue
    # check if in additional lists
    if c_OfficialLists:
        # extract lists from string
        # split string
        if DebugFlg:
            print(r, "List found")
        l_list = c_OfficialLists.split(';')
        i=0
        for tag in l_list:
            # look for tag in c_AdditionalInfo and extract string
            # for fraud, if HSS is found, tag a CORRECT HSS and no further processing
            regex = '\['+tag+'\].*?\['
            #_DEBUG print (regex)
            p = re.compile(regex)
            x = p.search(c_AdditionalInfo)
            if x:
                TagStr.append(x.group())
                 # we do not need to strip the brackets
                print(r, "List match ", tag, "found            ", end="\r")
                i += 1
                Extra = True
            # end if
    #   # end for
        if DebugFlg:
            print(TagStr)
    # end if (lists)
    # we now have TagInfo populated
    # len(TagInfo) = mumber of elements (>0) or i , Ectra = True as flag, and i as the 
    # for str in TagInfo:
    #  we check crimes and convictions there as well at the end ofr the following loop
    # check for convvicted crimes in Report
    sic_crime = check_issue(crimes, c_Triage, r)
    # now we check for Addidional List Tags
    if sic_crime == True:
        continue # go to next record
    if Extra:
        print("Checking additional in Lists", end="\r")
        LongReport = False
        ListCheck = True
        for x_Triage in TagStr:
            if len(x_Triage) > 720 and not CheckTag:
                LongReport = True
                ws.cell(row=r, column=35, value="LONG LIST ENTRY: REVIEW")
                print(r, "Long list enry.                         ", end='\r')
                continue
            sic_crime = check_issue(crimes, x_Triage, r)
            if sic_crime == True:
                break
        # end for
    # end if
    # if sic_crime was true, it was already written
    if sic_crime == True:
        continue
    if sic_crime == False:
        if preconv_option:
            print('SIC incorrect                                   ', end='\r')
            if CheckTag:
                pass # leave TAG SHOULD BE REMOVED
            else:
                ws.cell(row=r, column=35, value="INCORRECT")
            continue
        if not LongReport:
            print(r, "SIC incorrect                              ", end='\r')
            if pre_conv == False:
                if CheckTag:
                    pass # leave tag in status
                else:
                    ws.cell(row=r, column=35, value="INCORRECT")
            else:
                if CheckTag:
                    ws.cell(row=r, column=35, value="TAG SHOULD BE REMOVED (NO CONV)")
                else:
                    ws.cell(row=r, column=35, value="INCORRECT NO CONV (REVIEW)")
    if sic_crime == None:
        print(r, "Review manually                            ", end='\r')
        ws.cell(row=r, column=35, value="REVIEW MANUALLY")
    
    # end loop through rows
# write to new workbook
if not CheckTag:
    dws = wb['Tag should be removed']
else:
    dws = wb['Rev Manually blank no narc tria']
# delete the other sheet
print('\nRemoving other sheet ')
wb.remove(dws)
print('\nWriting and saving results in file', dest_file, '...' )
try:
    wb.save(dest_file)
except:
    input("\nCannot write to file. Try to close it first and press enter > ")
    print("Saving...")
    wb.save(dest_file)
print('Done')
# end program ######################################################################################################
 