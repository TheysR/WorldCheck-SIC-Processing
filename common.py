#######################################################################
# common functions/ classes for WordlCheck Triages
# ExcelHeader: readdsadrs header and assigns column numbers
#######################################################################
from sre_compile import isstring
from openpyxl import workbook, load_workbook
import sys, re
import argparse
import datetime, os
##########################################################################
class ExcelHeader:
    """reads header row from excel worksheet and assigns column numbers."""
# called with ws (worksheet class)
# Init: assigns column numbers to the headers of a SIC file, using
# label as index. Adds Remarks label in first row at the end.
# Properties:
# col (type dict): col['Header_Name'] containes column number
# last_col: contains last valid column
# values are assigned at initialisation
# Methods:
# AddColumn(): adds a column header with specified name
# ver 1.0 : read headers and add status and remarks if not found
# ver 1.1 : added last_column property
# ver 1.2 : added Addcolumn method (not tested)
# ver 1.3 : added ProfilType to TYPE match in headers (found in disqualified)
########################################################################## 
    def __init__(self, ws):
        self.col = {
            'AdditionalInfo' : 0,
            'Bio' : 0,
            'Categories' : 0,
            'OfficialLists' : 0,
            'Remarks' : 0,
            'Reports' : 0,
            'Status' : 0,
            'Type' : 0
            }
        # worksheet is pointer to open ws
        if not ws:
            #  error
            print('ExcelHeader() : ws: No open or valid workheet')
            sys.exit()
        # read first line
        c = 0
        for c_col in ws.columns:
            c += 1
            match (ws.cell(row=1, column=c).value):
                case "Categories":
                    self.col['Categories'] = c
                case "BIO":
                    self.col['Bio'] = c
                case "AdditionalInformation":
                    self.col['AdditionalInfo'] = c
                case "OfficialLists":
                    self.col['OfficialLists'] = c
                case "REPORTS":
                    self.col['Reports'] = c
                case "TYPE" | "ProfileType":
                    self.col['Type'] = c
                case "STATUS" | "Status":
                    self.col['Status'] = c
        # report and additional info column must be present to process file
        if self.col['Reports'] == 0:
            print('ERROR: Reports column missing, check file')
            sys.exit()
        if self.col['AdditionalInfo'] == 0:
            print('ERROR: AdditionalInformation column missing, check file')
            sys.exit()
        if self.col['Categories'] == 0:
            print('ERROR: Caterories column missing, check file')
            sys.exit()
        # if Status column was not found, add it
        if self.col['Status'] == 0:
            c += 1
            self.col['Status'] = c
            ws.cell(row=1, column=self.col['Status'], value='STATUS') # write missing status column
        # next free column, add Remarks column
        c += 1
        self.col['Remarks'] = c
        c += 1
        ws.cell(row=1, column=self.col['Remarks'], value = 'REMARKS')
        if self.col['OfficialLists'] == 0:
            # we need to assing a value to column otherwise caller will fail. This will happen in control+regulations nolist    
            self.col['OfficialLists'] = 100
        # if 'Type' column does not exist, we create element as next column so col references do not raise errors.
        # 'Type' is not mandatory for processing
        if self.col['Type'] == 0:
            c +=1
            print('WARNING: Type column missing')
            self.missing_col = 'Type'
            self.col['Type'] = c
        else:
            self.missing_col = ''
        self.last_column = c
    # end _init
    
    def AddColumn(self, ws, label):
        ''' Adds a user defined column to the open worksheet '''
        # not tested
        if not isstring(label):
            print ('AddColumn(): wrong type label :', label, 'argument must be a string')
            sys. exit() # fatal error
        if not ws:
            #  error
            print('AddColumn(): No open or valid workheet')
            sys.exit()            
        self.last_column +=1  
        dct = { label : self.last_column }
        self.col.extend(dct)
        ws.cell(row=1, column=self.last_column, value = label)
        return self.last_column
        
# end class ExcelHeader
#############################################################################################
class ExcelFile:
    ''' Opens excel file (workbook and workheet) from command line arguments
        params: program name and version
        Rules for filenames and worksheets (from command lines arguments):
        - filename is org_File.xlsx. Argument can be specifided with or without .xlsx ending
        - Destination file is 'org_file Passed.xlsx'
        - Worksheet name is same as filename (without .xlsx), unless specified by -ws argument
        It is up to the calling program to honour the flags (pc, debug, test)  
    '''
#   Methods:
#   ExcelSave(): saves new workbook with assigned filename at __init__
#   Propetries:
#   org_file : file ro read
#   dest_file : file wo write (destination file)
#   worksheet : worksheet name
#   wb : open workbook object
#   ws : open worksheet object
#   debug_flag, preconv_option, test_flag : boolean flags for verbose, precon processing only
#       and testing
#   row_limit : max number of rows to process with test flag
#   several properties for statistics (at end of __init__). It is up to the calling program to
#   increment the aprropriate values

    def __init__(self, program, ver):
        parser = argparse.ArgumentParser(description='Run SIC File' , prog=program)
        parser.add_argument("--pc", help="Chcek pre-conviction only", action='store_true')
        parser.add_argument("--version",help="Displays version only", action='version', version='%(prog)s ' + ver)
        parser.add_argument("--debug", help="Debug mode (verbose)", action='store_true')
        parser.add_argument('filename', help="filename to read")
        parser.add_argument('-t', '--test', help='run for a limited number of rows', type=int)
        parser.add_argument('-ws', '--worksheet', help='worksheet name if different from filename', dest='wsheet')
        parser.add_argument('--nolist', help='option without official lists', action='store_true')
        args = parser.parse_args()
        if args.debug:
            print("Debug mode")
            self.debug_flag = True
        else:
            self.debug_flag = False
        if args.pc:
            print("Pre Conv mode")
            self.preconv_option = True
        else:
            self.preconv_option = False
        org_file = args.filename
        if args.test:
            self.test_flag = True
            self.row_limit = args.test
            print("Test mode: processing", self.row_limit, "rows")
        else:
            self.test_flag = False
            self.row_limit = 0
        self.nolist = False
        if args.nolist:
            self.nolist = True
        if ".xlsx" not in org_file:
            if self.preconv_option:
                self.dest_file = org_file + ' Preconv Passed.xlsx'
            else:
                self.dest_file = org_file + ' Passed.xlsx'
            ws_name = org_file
            self.org_file = org_file + '.xlsx'
        else:
            file_parts = org_file.split('.')
            if self.debug_flag:
                print(file_parts)
            ws_name = file_parts[0]
            if self.preconv_option:
                self.dest_file = file_parts[0] + ' Preconv Passed.xlxs'
            else:
                self.dest_file = file_parts[0] + ' Passed.xlxs'
            self.org_file = org_file
        # override worksheet name if specified
        if args.wsheet:
            self.worksheet = args.wsheet
        else:
            self.worksheet = ws_name

        # open workbook

        print( 'Loading spreadsheet ', self.org_file)
        # check if filename exists
        #
        try:     
            self.wb = load_workbook(filename=self.org_file)
        except:
            print("cannot open file ", self.org_file)
            sys.exit()
        try:
            self.ws = self.wb[self.worksheet]
        except:
            print("cannot open worksheet ", self.worksheet)
            sys.exit()
        self.long_entries = 0
        self.preconv = 0
        self.postconv = 0
        self.entities = 0
        self.other = 0
        self.off_lists = 0
        self.review = 0
        self.sic_correct = 0
        self.sic_incorrect = 0
        self.no_report = 0
    # end __init__

    def ExcelSave(self):
        ''' Save workbook '''
        try:
            self.wb.save(self.dest_file)
        except:
            input("\nCannot write to file. Try to close it first and press enter > ")
            print("Saving...")
            self.wb.save(self.dest_file)
# end class ExcelFile
############################################################################################
def RegexSearch(regex, String, r):
    """"Search for string in text using regex."""
# helper function to search for regular expression to catch error
# and use case ignore flag
#############################################################################################
    try:
        p = re.compile(regex, re.IGNORECASE)
    except:
        print(r, 'Wrong regex:', regex)
        sys.exit()
    mtch = p.search(String)
    return mtch
# end function
#############################################################################################
class Logger:
    '''Writes to log file '''
    def __init__(self, outfile):
        if ".log" not in outfile:
            self.dest_file = outfile + '.log'
        else:
            self.dest_file = outfile
        t = datetime.datetime.now()
        
        try:
            f = open(self.dest_file, 'a')
        except:
            print('Cannot open log file ',self.dest_file)
            return False
        self.stream = f

        print('\n', t, file=f)
        print('----------------------------', file=f)
    
    # write without timestamp    
    def output(self, txt):        
        try:
            print(txt, file=self.stream)
            return True
        except:
            print('Cannot write to file ', self.dest_file)
            return False
    
    # write timestamp    
    def timestamp(self):
        print(datetime.datetime.now(), file=self.stream)
        return True
    
    # write with timestamp    
    def toutput(self, txt):
        print(datetime.datetime.now(),': ', txt, file=self.stream)
        return True
    
    # method to expliclty close stream (should no be used often)    
    def close(self):
        os.close(self.stream)
