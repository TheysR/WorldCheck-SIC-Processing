#######################################################################
# parse Excel Worksheet for correct SIC tag
# LOGIC FOR CONGTROL AND REGULATION
# (c) 2022 Theys Radmann
# ver 1.0
# ver 2.0 : Entities is prcvessed, added Financial Services Warnings positive
# This program introduces statistics and logging, and should be used as template 
# for futures SIC programs. Uncomment sections as necessary
#######################################################################
# modules/libararies needed
from openpyxl import load_workbook, Workbook
import re  # regex
import sys
import argparse
from common import ExcelFile, ExcelHeader, RegexSearch, Logger

# definition of offense categories
# version
ver = '2.1'
# log file name and program name
logfile = 'control_reg'
program_name = 'control_regulations'

# offense triage (regular expressions). 
# order of some crimes in list is important for logic and efficiency
# First crime found and convicted for, ends checks for further crimes. Therefore,
# most frequent ones should ideally be first 
crimes = [
    r"anti-?trust violations?",
    r"banking regulation violations?",
    r"breach of AML regulations?",
    r"breach of market violation law",
    r"serious disciplinary and law violations",
    r"(infringment of|contravening) ((antimafia|building|civil service))? regulations?",
    r"contravening parallel( and direct) import regulations?",
    r"contravening the (clean air|exchange control) regulations?"
    r"failing to comply with",
    r"highway traffic act violations",
    r"infraction of environmental laws and regulations",
    r"(infraction|infringment) of( the)?( central)? bank regulations?",
    r"law violations",
    r"infrignments of securities legislation and regulations",
    r"non-compliance with regulations",
    r"regulation violation",
    r"voilated the (banking act|local tax law)",
    r"violated the public office election (act|law)",
    r"violated the Unfair Competition Prevention Act",
    r"violated Unauthorised Computer Access Law",
    r"violating( the)? ((copyright( infringment)?)|immigration|intellectual property) (laws?|act)",
    r"violation of the phillipine intellectual property code",
    r"violating environmental protection regulations",
    r"violating insurance laws",
    r"violating regulations on bidding",
    r"(violating|violation of) the (banking|civil service|clean air|dodd-frank|foreign business|securities( and excange)?|wildlife) Act",
    r"(violation of|violating) the (truth in lending|argicultural finance) act"
    r"violating the( colorado)? organi[sz]ed crime control act",
    r"violating the( \w+?)? securities act",
    r"(violating|violation of) the (anti-smuglging|public office election) law",
    r"violating( new york)? sate banking law",
    r"violating the law on financing of terrorism",
    r"violating the Racketeer Influenced and Corrupt Organi[sz]ations Act",
    r"violations? against market law",
    r"violations? of (anti moneylaundering|anti-graft practices) act",
    r"violations? of( the)? copyrights? (act|law)",
    r"violations? of the duty",
    r"violations? of the( prevention of)? money laundering( control)? act"
    r"violations? of (loterry regulation|political funds|the drug|the martin|anti-corruption|antimonopoly|clean water) act",
    r"violations? of the (investment|customs?|banking|air pollution|futures trading|forest|financial instrument and exchange) act",
    r"violations? of the (trademark|(wildlife( protection?))|trade secrets|marketing|moneylenders|mineral resources|national forestry) act",
    r"violations? of the (organized crime prevention|securities and exchange|waste disposal) act",
    r"violations? of the (sexual offences against children|wildlife resources conservation and protection) act",
    r"violations? of the (employmnent service|telecommunucations|trade secrets|organi[zs]ed crime prevention) act",
    r"violations? of the (Public Procurement and Disposal|street crime and terrorism prevention) act,"
    r"violations? of the export administration ragulations",
    r"violations? of Financial Investment Services and Capital Market Act",
    r"violations? of the food and drugs?( administration)? act",
    r"violations? of the Code of Professional Responsibility",
    r"violations? of the anti-trafficking",
    r"violations? of the Anti-Graft and Corrupt Practices Act",
    r"violations? of the Government Procurement Act",
    r"violations? of the Revised Philippine Forestry Code",
    r"violations? of the( local tax)? law",
    r"violations? of (securities|electoral|anti-narcotics|illegal immigration|against market|consumer) law",
    r"violations? of the competition protection law",
    r"violations? of the Export Administration Regulations?",
    r"(infringment|violations?) of( significant|the)? industrial (property|protection) (rights|laws)",
    r"(breach|violations?) of the consumer (loan|protection) act",
    r"violations? of( the)? law",
    r"customs violation",
    r"violations? of Prohibition of Fraudulent and Unfair Trade Practices Regulations",
    r"violations? of regulations of firearms",
    r"violations? of information disclosure regulations?",
    r"(infraction|violations?) of environmental laws?",
    r"violations? of Securities( and) Exchange Act",
    r"violations? of( the) intellectual propetry code",
    r"violation of human trafficking",
    r"violation of banking regulations?",
    r"violations? of the Offences and Penalties Regulation",
    r"violations under the Foreign Exchange Management Act",
    r"violation of professional secrecy",
    r"violations? of the Act Governing Food",
    r"violations? of the Act on Specified Commercial Transactions",
    r"offences in accordance to the Law No 29622 and its regulations",
    r"breaches of the Act relating to supervision deficiencies",
    r"violation of Prohibition of Fraudulent and Unfair Trade Practices",
    r"(infringing|violations? of) the competition act",
    r"lack of organisation and internal control",
    r"(violating|violation(s) of)( bank)? lending regulations",
    r"violating regulations on( exploitation and)? protection of (forests|endangered and rare animals)",
    r"violation of the regulations governing environmental protection",
    r"violating( state)? regulations on economic management",
    r"violating regulations on research, exploration and exploitation of natural resources",
    r"breaching( Alberta's)? fish and wildlife regulations",
    r"violating regulations on management and use of state assets",
    r"contravening the Marine Living Resources Act",
    r"serious infringement of rules and regulations",
    # additional (with lists only)
    r" not complying with the requirements of international maritime regulations",
    r"violation of legislation on protection of economic competition",
    r"violated the anti-money laundering provisions of the Bank Secrecy Act",
    r"violation of the International Emergency Economic Powers Act",
    r"violation of financial legislation",
    r"violation of the Colorado Securities Act",
    r"violations of the Clean Air Act",
    r"violations of regulations in relation to Law",
    r"Medicaid Claims Act violations",
    r"violated the Controlled Substances Act",

]
# below acquittal/dismissal triages should not be changed, unless new insight is found
aquittals = [
    r"ac?quitt(al|ed)",
    r"pardon(ed)?",
    r"case filed",
    r"dismissed",
    r"dropped"
]
dismissals = [
    r"dismiss(ed|al)",
    r"dropped",
    r"case filed"
]
# modify as appropriate. Values 20 and 800 are directed as satisfactory thesholds
words_apart = 20 # maximum distance of words apart from crime and conviction when matching offence frist and conviction second
max_rep_length = 800 # maximum report length for processing, longer that this will get tagged for review

# initialise some global variables
pre_conv = False
Debug_Flg = False
Test_Flag = False
True_Condition = False
offence_found = False
sic_tag = False
has_list = False

# functions

####################################################################
def check_list_sic(list_tag, r):
    ''' Checks for list that have always SIC CORRECT ststus'''
#  returns true or false for trapping positive sic lists
####################################################################
# lists that trigger positive sic tag (could be read from a file)
# not used in this program (control_reg)
    
    true_lists =[] # put list names in here
    list_status = [ False, "Null"]
    for str_list in true_lists:
        if str_list in list_tag:
            list_status = [ True , str_list]
            return list_status
    return list_status 
# end check_sic_list()
############################################################
def check_conviction(type, str_report, n):
    ''' Check if there was a convitcion for the offence type'''

# type : crime/offence (string/regex)
# str_report : record (report column) (string)
# r: row begin processed, for informational purposes only (debugging)
# returns 1 if found and issue follows conviction (no writes)
# returns 2 if found and issue is followed by conviction (no writes)
# returns -1 if issue is followed by conviction but too far apart (writes)
# returns -2 if conviction found with reversal (acquittal) found (no writes)
# returns 0 is no conviction was found at all (no writes)
# the reason for write case (-1) is in case there is no further processing that changes
# the status later. 
############################################################
    post_conv = 0
    long_flag = False
    global words_apart, Debug_Flg
    phrase = [
        r"found guilty",
        r"convicted",
        r"sentence[d]?",
        r"pleaded guilty",
        r"pleaded no contest",
        r"imprisoned",
        r"fined",
        r"arrested .+ serve",
        r"for conviction",
        r"ordered .*\s*to (pay|serve)",
        r"incarcerated",
        r"admitted guilt",
        r"served probation",
        r"to serve .* imprisonment",
        r'previous conviction[s]* .*?'
    ]
    # keywords must be near crime type if before conv
    # build search string with crime type
   
    for str in phrase:
        long_flag = False
        # search conviction before crime. Distance of words are checked here, but may not be necessary, 
        # as crime usually follows conviction after a few words if specified after.
        s_str = str + ' .*?' + type  # RegEx word followed by space and anythnig in between and the second word
        x = RegexSearch(s_str, str_report, n)
        if x: 
            # conviction found
            if Debug_Flg:
                log.output(n, x.group())
            # check if distance between offense and convictions is large
            words = re.split('\s', x.group())
            if len(words) > words_apart:
                # crime too far from sentence, look for another sentence further ahead, in case there are two
                n_idx = slice(x.start()+1, len(str_report)-1)
                y =RegexSearch(s_str, str_report[n_idx], n)
                if y:
                    words = re.split("\s", y.group())
                    if len(words) > words_apart:
                        print (n, "Too many words between issue and conv", end="\r")
                        post_conv = -1 # to flag for review
                        print(r, "SIC Review                            ", end='\r')
                        ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
                        if ListCheck:
                            ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: Remote connection between conviction and offence. From List")
                            xls.review += 1
                        else:
                            ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: Remote connection between conviction and offence")
                            xls.review += 1
                        return -1  # issue is too far for a conclusive conviction
                    # end if len()
                # enf if y
            # end if (len)
            # we have foud conviction, check for aquittals
            for tag in aquittals:
                s_str = type + '.*' + tag
                s_aquitt = RegexSearch(s_str, str_report, n)
                if s_aquitt:
                    print(n, "Dismissal found                                ", end='\r')   
                    return -2
                    # this may be revised
            # end for (aquitals)
            return 2
            
        # if not found, check the other way around (conviction after crime). 
        # Problem is, if there was a conviction for something different, in which
        # case we should not check for preious mentions of issues
        # this is difficult. Here some tries just to catch these common ones
        if "sentenced for" in str_report:
            continue
        if "pleaded guilty to" in str_report:
            continue
        if "found guilty for" in str_report:
            continue
        if "pleaded no contest to" in str_report:
            continue
        s_str = 'sentenced for \d+ years'
        x = RegexSearch(s_str, str_report, n)
        if not x:
            s_str = "sentence[d]* .*? *for "
            x = RegexSearch(s_str, str_report, n)
            if x:
                continue
        else:
            if Debug_Flg:
                print(type, "\n", str)
                print(n, "sentenced for x years found")
                input("enter")
        s_str = "sentence[d]* .*? *on charges of"
        x= RegexSearch(s_str, str_report, n)
        if x:
            continue
        s_str = "found guilty .*? *on charges of"
        x= RegexSearch(s_str, str_report, n)
        if x:
            continue
        s_str = 'pleaded guilty .*? *to'
        x= RegexSearch(s_str, str_report, n)
        if x:
            continue
        # now the check conviction after crime
        s_str = type + r'.*? ' + str
        x = RegexSearch(s_str, str_report, n)
        if x:
            if Debug_Flg:
                print(n, x.group())
            # found. if there are too many words between the type and the conviction phrase, assume review
            # there may be a later repetiion of the word (crime) wich decreases the word count, we check twice
            words = re.split("\s", x.group()) # split into words

            if len(words) > words_apart: # too many words in between, but there could be further mention of issue
                # look for issue further ahead
                n_idx = slice(x.start()+1, len(str_report)-1)
                y = RegexSearch(s_str, str_report[n_idx], n)
                if y:
                    words = re.split("\s", y.group())
                    if len(words) > words_apart:
                        print (n, "Too many words between issue and conv", end="\r")
                        post_conv = -1 # to flag for review
                        long_flag = True
                # end if (y)
            # end if (Len)
            if long_flag == False:
                for tag in aquittals:
                    s_str = type + '.*' + tag
                    s_aquitt = RegexSearch(s_str, str_report, n)
                    if s_aquitt:
                        print(n, "Aquittal found                                ", end='\r')   
                        return -2 
                    # this may be revised
                # end for (aquitals)
            # end if (long_flag)
            # found without aquittal
            return 1
        # end if (x)
    # end for (str)
    return post_conv
#####################################################################
def check_item(item, str_Triage, r, pre_conv, src_text):
    '''Checks triage for record and writes approriate tag if offence/issue was found. '''
# normally called by check_issues()
# will not write to record if no match was found or exception (long content)
#####################################################################
    global Debug_Flg, ListCheck, ws, dismissals
    sic_tag = False
    s_crime = RegexSearch(item, str_Triage, r)
    if s_crime:
        if pre_conv:
            # check for dismissed
            for tag in dismissals:
                s_str = item + '.*' + tag # we may ommit item in search string
                s_diss = RegexSearch(s_str, str_Triage, r)
                if s_diss:
                    print(r, "Dismissal found                                ", end='\r')   
                    ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
                    ws.cell(row=r, column=head.col['Remarks'], value="Pre Conv: Dismissal found ["+src_text+"]")
                    xls.review += 1
                    return True
                # end if
            # end for
            print(r, 'SIC correct                                            ', end='\r')
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
            ws.cell(row=r, column=head.col['Remarks'], value="Pre Conv ["+src_text+"]")
            xls.sic_correct += 1
            return True
        # end if preconv
        # post conv processing
        chk = check_conviction(item, str_Triage, r)
        if chk == -1:
            # too far away, flag for review (for now). It was written in file in function
            print(r, "SIC Review                     ", end='\r')
            sic_tag = None
        if chk == 1:
            # write correct to sheet
            print(r, "SIC Correct                             ", end='\r')
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
            ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: Conviction after Triage ["+src_text+"]")
            xls.sic_correct += 1
            return True
        if chk == 2:
            # write correct to sheet
            print(r, "SIC Correct                             ", end='\r')
            ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
            ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: Conviction before Triage ["+src_text+"]")
            xls.sic_correct += 1
            return True
        if chk == -2:
            print(r, 'Review manually                                ', end='\r')
            ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
            ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: acquittal found ["+src_text+"]")
            xls.review += 1
            return True
        # if no conviction was found, check if there was a dismissal
        if chk == 0:
            for tag in dismissals:
                s_str = item + '.*' + tag # we may ommit item in search string
                s_diss = RegexSearch(s_str, str_Triage, r)
                if s_diss:
                    print(r, "Dismissal found                                ", end='\r')   
                    ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
                    ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: No conviction, dismissal found ["+src_text+"]")
                    xls.review += 1
                    return True # behaves like correct as no further offences are affeced if there is a dismissal
                    # this may be revised
                # end if
            # end for
            print (r, 'Tag correct - pre conv', end='\r')
            ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
            ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: No conviction ["+src_text+"]")
            xls.review += 1
            return True

    # end if (s_crime true)
    return sic_tag  # can only be False or None here
        

#####################################################################
def check_issues(issues, str_Triage, r, preconv, Source):
# checks if crime was found and convicted
# # returns True if offense found
# (crime found and written in record in check_item()), 
# False (no tag/offense found), 
# and None (to review, long list or report, or distance) 
# 
#####################################################################
    global Debug_Flg, ws
    sic_tag = False
    # first, let's check review tags

    for x_crime in issues:
        sic_tag = check_item(x_crime, str_Triage, r, preconv, Source)
        if sic_tag:
            break
    # end for (issues loop)
    return sic_tag
# end functions
###################################################
# Main
# start program
# here two files are read, one with official lists and on without.
# no officiallists, read from reports
# with officialist, read additional info
# the option wil be given by argument to program
Test_Flag = False
Debug_Flg = False
True_Condition = False
offence_found = False
row_limit = 0
sic_tag = False
has_list = False

# opem logger

log = Logger(logfile)

xls = ExcelFile(program_name, ver)

if xls.debug_flag:
    Debug_Flg = True
if xls.test_flag:
    row_limit = xls.row_limit
    Test_Flag = True
if xls.nolist:
    no_list = True
else:
    no_list = False    
ws = xls.ws
head = ExcelHeader(ws)

if Debug_Flg:
    print(head.col)
    input('Enter > ')
# for each row
r = 0
for row in ws.rows:
    pre_conv = False
    r += 1
    if r == 1:
        continue
    if Test_Flag:
        if r > row_limit:
            break
    # read row into variables (only useful ones)
    c_categories = ws.cell(row=r, column=head.col['Categories']).value       
    c_OfficialLists = ws.cell(row=r, column=head.col['OfficialLists']).value    
    c_AdditionalInfo = ws.cell(row=r,column=head.col['AdditionalInfo']).value
    c_Reports = ws.cell(row=r,column=head.col['Reports']).value     
    c_Type = ws.cell(row=r, column=head.col['Type']).value
    # c_Bio = ws.cell(row=r, column=head.col['Bio']).value
    c_status= ws.cell(row=r, column=head.col['Status']).value
    # accorging to this sic, option no list reads from reports, otherwise from additionalinfo
    if no_list:
        c_Triage = c_Reports
        c_source = "RPT"
    else:
        c_Triage = c_AdditionalInfo
        c_source = "LST"
    c_lists = [] # resets list of OifficialLists
    has_list = False
    Long_Report = False
    ListCheck = False
    sic_tag = False
    if head.missing_col == 'Type':
        # classify according to caterogy (control/reg)
        match c_categories:
            case "VESSEL" | "EMBARGO VESSEL" | "CORPORATE" | "ORGANISATION" | "POLITICAL PARTY" | "TRADE UNION" | "PORT" | "BANK":
                c_Type = "E"
                xls.entities += 1
            case _: 
                c_Type = "I"
   
    # uptate remarks column to be able to filter for entities (control/reg)
    if c_Type == 'E':
        ws.cell(row=r, column=head.col['Categories'], value='ENTITY:' + c_categories)
        
    if not c_Reports and no_list: # no_List option take triage from report in control/reg
        ws.cell(row=r, column=head.col['Status'], value='NO REPORT')
        ws.cell(row=r, column=head.col['Remarks'], value='No report column')
        print(r, "No report found.                         ", end='\r')
        xls.no_report +=1
        continue
    if no_list == False:
        if len(c_Reports) > max_rep_length:
            ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
            ws.cell(row=r, column=head.col['Remarks'], value='LONG CONTENT')
            print(r, "Report too long.                         ", end='\r')
            xls.long_entries +=1
            xls.review += 1
            continue
    # no true lists in control+reg
    #if c_OfficialLists and c_OfficialLists != "NULL":
    #    sic_list = check_list_sic(c_OfficialLists, r)
    #    if sic_list[0] == True:
    #        print(r, "Tagged list", end='\r')
    #        ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
    #        ws.cell(row=r, column=head.col['Remarks'], value='OFFICIAL LIST : ' + sic_list[1])
    #        xls.off_lists += 1
    #        xls.sic_correct += 1
    #        continue # no further processing needed
#    if c_OfficialLists and c_OfficialLists != "NULL":
#        if Debug_Flg:
#            print(r, "List found")
#        l_list = c_OfficialLists.split(';')
#        i=0
#        for tag in l_list:
#            # look for tag in c_AdditionalInfo and extract string
#            regex = '\['+tag+'\].*?\['
#            #_DEBUG print (regex)
#            x = RegexSearch(regex, c_AdditionalInfo, r)
#            if x:
#                c_lists.append(x.group())
#                 # we do not need to strip the brackets
#                print(r, "List match ", tag, "found            ", end="\r")
#                i += 1
#                has_list = True
#            # end if
#        # end for
#    # end if (OfficialLists)
 #   if Debug_Flg:
 #       print(c_lists)
    
    # we now have TagInfo populated
    # Review keywords
    if "[FINANCIAL SERVICES WARNINGS]" in c_AdditionalInfo:
        ws.cell(row=r, column=head.col['Status'], value="SIC TAG CORRECT")
        ws.cell(row=r, column=head.col['Remarks'], value='FSW')
        print(r, "Financial Services Warning.                         ", end='\r')
        xls.sic_correct +=1
        xls.other += 1
        continue    

    # flag pre/post conv
    if "CRIME" in c_categories:
        pre_conv = False
        xls.postconv +=1
    else:
        pre_conv = True
        xls.preconv +=1
    sic_tag = check_issues(crimes, c_Triage, r, pre_conv, c_source)
    if sic_tag == True:
        continue # go to next record, ignore lists as a SIC tag was found
    # check in lists, never executed in control+regulations
   
    if has_list:
        print("Checking additional in Lists", end="\r")
        ListCheck = True
        for x_Triage in c_lists:
            Long_Report = False
            if len(x_Triage) > max_rep_length:
                Long_Report = True
                ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
                ws.cell(row=r, column=head.col['Remarks'], value="LONG REPORT [LIST]")
                print(r, "Long list enry.                         ", end='\r')
                xls.long_entries += 1
                xls.review += 1
                continue # next list entry
            if sic_tag:
                break # no more list cheks
            sic_tag = check_issues(crimes, x_Triage, r, pre_conv, 'LST')
            if sic_tag == True:
                break
        # end for (list)
    # end if (extra/ lists)
    # if sic_tag was true, it was already written
    if sic_tag == True:
        continue
    if sic_tag == False:
        if Long_Report:
            print (r, "Review manually.                                 ", end='\r')
            ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
            ws.cell(row=r, column=head.col['Remarks'], value="Content to long.")
            xls.long_entries += 1
            xls.review += 1
            continue
        if pre_conv:
            print(r, 'SIC incorrect                                   ', end='\r')
            ws.cell(row=r, column=head.col['Status'], value="TAG SHOULD BE REMOVED")
            ws.cell(row=r, column=head.col['Remarks'], value="Pre Conv: No offence found")
            xls.sic_incorrect += 1
            continue
        print(r, 'SIC incorrect                                   ', end='\r')
        ws.cell(row=r, column=head.col['Status'], value="TAG SHOULD BE REMOVED")
        ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: No offence found")
        xls.sic_incorrect += 1
    if sic_tag == None:
        print(r, "Review manually                            ", end='\r')
        ws.cell(row=r, column=head.col['Status'], value="REVIEW MANUALLY")
        ws.cell(row=r, column=head.col['Remarks'], value="Post Conv: relation between crime and conviction not clear")
        xls.review += 1
    # end loop through rows
# write to new workbook
print('\nWriting and saving results in file', xls.dest_file, '...' )
try:
    xls.ExcelSave()
except:
    input("\nCannot write to file. Try to close it first and press enter > ")
    print("Saving...")
    xls.ExcelSave()
# print summary. perhaps write to log fle?
print('Done')
log.toutput('Done')
print('Summary')
log.output('Summary')
print('=======')
log.output('=======')
print('Entities:\t',xls.entities)
log.output('Entities:\t',xls.entities)
print('Long Entries:\t',xls.long_entries)
log.output('Long Entries:\t',xls.long_entries)
print('Official Lists:\t', xls.off_lists)
log.output('Official Lists:\t', xls.off_lists)
print('No Report:\t',xls.no_report)
log.output('No Report:\t',xls.no_report)
print('Pre Conv:\t', xls.preconv)
log.output('Pre Conv:\t', xls.preconv)
print('Post Conv:\t',xls.postconv)
log.output('Post Conv:\t',xls.postconv)
print('Other:\t\t', xls.other)
log.output('Other:\t\t', xls.other)
print('SIC Correct:\t',xls.sic_correct)
log.output('SIC Correct:\t',xls.sic_correct)
print('SIC Incorrect:\t',xls.sic_incorrect)
log.output('SIC Incorrect:\t',xls.sic_incorrect)
print('Man. Review:\t',xls.review)
log.output('Man. Review:\t',xls.review)
print('Total:\t\t', r)
log.output('Total:\t\t', r)
# end program ######################################################################################################
 